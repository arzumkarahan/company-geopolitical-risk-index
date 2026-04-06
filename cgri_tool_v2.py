"""
Corporate Geopolitical Risk Index (CGRI) — Interactive Tool  v2
================================================================
Methodology strictly follows the Excel template in data/processed/.

Formula
-------
CGRI = (0.15 × HQ_Risk
       + 0.45 × Revenue_Exposure
       + 0.40 × SupplyChain_Exposure)
       × Sector_Multiplier
       × Volatility_Multiplier
       × Financial_Leverage_Multiplier

Component formulas
------------------
HQ Risk                    : country GRI of headquarters country
Revenue Exposure           : Σ(GRI_c × rev_std_c) × HHI_sub(Σ rev_std_c²)
Supply Chain               : (0.5 × sup_comp + 0.5 × fac_comp) × HHI_sub(avg_HHI)
Financial Leverage Mult.   : 0.8 (Net D/EBITDA < 0), 0.9 (0–2), 1.0 (2–4), 1.1 (4–6), 1.2 (≥ 6)
"""

from pathlib import Path
import json

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl import load_workbook

st.set_page_config(
    page_title="CGRI Tool",
    layout="wide",
    page_icon="🌐",
    initial_sidebar_state="expanded",
)

# ── Custom CSS ───────────────────────────────────────────────────────────────
st.markdown("""
<style>
/* ── Global ── */
[data-testid="stAppViewContainer"] { background: #f7f8fc; }
[data-testid="stSidebar"] { background: #1a1d2e; }
[data-testid="stSidebar"] * { color: #d0d4e8 !important; }
[data-testid="stSidebar"] .stSelectbox label,
[data-testid="stSidebar"] .stNumberInput label,
[data-testid="stSidebar"] .stTextInput label { color: #9aa0c0 !important; font-size: 0.8rem; }
[data-testid="stSidebarContent"] h1,
[data-testid="stSidebarContent"] h2,
[data-testid="stSidebarContent"] h3 { color: #ffffff !important; }

/* ── KPI cards ── */
.kpi-card {
    background: #ffffff;
    border-radius: 12px;
    padding: 18px 20px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.07);
    border-left: 4px solid #4b6fff;
    margin-bottom: 4px;
}
.kpi-label { font-size: 0.75rem; color: #8a94a6; font-weight: 600; letter-spacing: 0.05em; text-transform: uppercase; }
.kpi-value { font-size: 1.9rem; font-weight: 700; color: #1a1d2e; line-height: 1.1; }
.kpi-sub   { font-size: 0.78rem; color: #a0a8b8; margin-top: 2px; }

/* ── Result badge ── */
.score-badge {
    border-radius: 14px;
    padding: 28px 30px;
    text-align: center;
    box-shadow: 0 4px 18px rgba(0,0,0,0.15);
}
.score-badge .score-num { font-size: 3.2rem; font-weight: 800; color: #fff; line-height: 1; }
.score-badge .score-cat { font-size: 1.1rem; color: rgba(255,255,255,0.88); margin-top: 6px; }
.score-badge .score-co  { font-size: 0.85rem; color: rgba(255,255,255,0.65); margin-top: 4px; }

/* ── Section header ── */
.section-hdr {
    font-size: 0.7rem; font-weight: 700; letter-spacing: 0.12em;
    text-transform: uppercase; color: #8a94a6; margin: 12px 0 6px;
}

/* ── Metric card (component) ── */
.comp-card {
    background: #fff;
    border-radius: 10px;
    padding: 14px 16px;
    box-shadow: 0 1px 5px rgba(0,0,0,0.06);
    text-align: center;
}
.comp-card .cv  { font-size: 1.55rem; font-weight: 700; color: #1a1d2e; }
.comp-card .cl  { font-size: 0.7rem; color: #8a94a6; font-weight: 600;
                  text-transform: uppercase; letter-spacing: 0.05em; }
.comp-card .cd  { font-size: 0.73rem; color: #b0b8c8; margin-top: 2px; }

/* ── Streamlit tabs tweak ── */
button[data-baseweb="tab"] { font-size: 0.95rem; font-weight: 600; }

/* ── Mobile responsive ── */
@media screen and (max-width: 768px) {
    /* Collapse the sidebar on mobile so it doesn't block the screen */
    [data-testid="stSidebar"] {
        transform: translateX(-110%) !important;
        visibility: hidden !important;
    }
    [data-testid="stSidebarCollapsedControl"] {
        visibility: visible !important;
    }

    /* Stack all multi-column layouts vertically */
    [data-testid="stHorizontalBlock"] {
        flex-wrap: wrap !important;
    }
    [data-testid="column"] {
        width: 100% !important;
        flex: 1 1 100% !important;
        min-width: 100% !important;
    }

    /* Tighter main content padding */
    .main .block-container {
        padding-left: 0.75rem !important;
        padding-right: 0.75rem !important;
        padding-top: 1rem !important;
    }

    /* KPI cards: reduce font size so they fit */
    .kpi-value { font-size: 1.4rem !important; }
    .kpi-card  { margin-bottom: 8px !important; }

    /* Component cards */
    .comp-card .cv { font-size: 1.2rem !important; }

    /* Score badge */
    .score-badge .score-num { font-size: 2.4rem !important; }

    /* Country input table: make the share % and delete button compact */
    [data-testid="stHorizontalBlock"] [data-testid="column"]:nth-child(3),
    [data-testid="stHorizontalBlock"] [data-testid="column"]:nth-child(4) {
        flex: 0 0 auto !important;
        width: auto !important;
        min-width: unset !important;
    }

    /* Inputs: bigger touch targets */
    .stSelectbox > div, .stNumberInput > div {
        min-height: 44px !important;
    }

    /* Charts: prevent horizontal overflow */
    .js-plotly-plot { max-width: 100% !important; }

    /* Tables: allow horizontal scroll instead of squishing */
    [data-testid="stDataFrame"] { overflow-x: auto !important; }

    /* Hide the wide-layout padding Streamlit adds */
    [data-testid="stAppViewContainer"] > section {
        padding-left: 0 !important;
        padding-right: 0 !important;
    }
}
</style>
""", unsafe_allow_html=True)

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data" / "processed"

# ═══════════════════════════════════════════════════════════════════════════
# CORE COMPUTATION FUNCTIONS  (mirror the Excel formulas exactly)
# ═══════════════════════════════════════════════════════════════════════════

def normalize_weights(exposures: dict) -> dict:
    cleaned = {
        str(k).strip(): float(v)
        for k, v in exposures.items()
        if k is not None and pd.notna(v) and float(v) > 0
    }
    total = sum(cleaned.values())
    return {k: v / total for k, v in cleaned.items()} if total > 0 else {}


def weighted_country_gri(shares: dict, country_risk: dict) -> float:
    missing = [c for c in shares if c not in country_risk]
    if missing:
        raise KeyError(f"Countries not in GRI dataset: {missing[:5]}")
    return sum(country_risk[c] * w for c, w in shares.items())


def hhi(shares: dict) -> float:
    return sum(w ** 2 for w in shares.values())


def hhi_submultiplier(h: float) -> float:
    if h < 0.15: return 0.90
    if h < 0.25: return 1.00
    if h < 0.40: return 1.10
    if h < 0.60: return 1.25
    return 1.50


def net_debt_to_financial_multiplier(ratio: float) -> float:
    x = float(ratio)
    if x < 0: return 0.8
    if x < 2: return 0.9
    if x < 4: return 1.0
    if x < 6: return 1.1
    return 1.2


def compute_revenue_exposure(revenue_by_country: dict, country_risk: dict) -> dict:
    std = normalize_weights(revenue_by_country)
    intermediate = weighted_country_gri(std, country_risk)
    rev_hhi = hhi(std)
    sub = hhi_submultiplier(rev_hhi)
    return {"std_shares": std, "intermediate": intermediate,
            "hhi": rev_hhi, "hhi_sub": sub, "final": intermediate * sub}


def compute_supply_chain_exposure(
    suppliers_by_country: dict,
    supplier_facilities_by_country: dict,
    country_risk: dict,
) -> dict:
    sup_std = normalize_weights(suppliers_by_country)
    fac_std = normalize_weights(supplier_facilities_by_country)
    sup_comp = weighted_country_gri(sup_std, country_risk)
    fac_comp = weighted_country_gri(fac_std, country_risk)
    intermediate = 0.5 * sup_comp + 0.5 * fac_comp
    hhi_sup = hhi(sup_std)
    hhi_fac = hhi(fac_std)
    hhi_combined = (hhi_sup + hhi_fac) / 2
    sub = hhi_submultiplier(hhi_combined)
    return {
        "sup_std": sup_std, "fac_std": fac_std,
        "sup_component": sup_comp, "fac_component": fac_comp,
        "intermediate": intermediate,
        "hhi_sup": hhi_sup, "hhi_fac": hhi_fac,
        "hhi_combined": hhi_combined, "hhi_sub": sub,
        "final": intermediate * sub,
    }


def compute_facility_risk(facility_by_country: dict, country_risk: dict) -> dict:
    std = normalize_weights(facility_by_country)
    component = weighted_country_gri(std, country_risk)
    return {"std_shares": std, "final": component}


def compute_cgri(
    hq_country: str, sector: str, net_debt_ebitda: float,
    revenue_by_country: dict, suppliers_by_country: dict,
    supplier_facilities_by_country: dict,
    country_risk: dict, sector_mult_lookup: dict,
    volatility_mult: float, company_name: str = "Custom Company",
) -> dict:
    hq_country = str(hq_country).strip()
    sector     = str(sector).strip()
    if hq_country not in country_risk:
        raise KeyError(f"HQ country '{hq_country}' not found in GRI dataset.")
    if sector not in sector_mult_lookup:
        raise KeyError(f"Sector '{sector}' not found in sector multiplier table.")

    hq_risk  = float(country_risk[hq_country])
    rev      = compute_revenue_exposure(revenue_by_country, country_risk)
    sc       = compute_supply_chain_exposure(suppliers_by_country, supplier_facilities_by_country, country_risk)
    fin_mult = net_debt_to_financial_multiplier(net_debt_ebitda)
    sec_m    = float(sector_mult_lookup[sector])

    base  = 0.15*hq_risk + 0.45*rev["final"] + 0.40*sc["final"]
    final = base * sec_m * volatility_mult * fin_mult

    return {
        "company": company_name, "hq_country": hq_country, "sector": sector,
        "net_debt_ebitda": float(net_debt_ebitda),
        "hq_risk": hq_risk, "revenue_exposure": rev["final"],
        "supply_chain": sc["final"],
        "financial_multiplier": fin_mult,
        "sector_multiplier": sec_m, "volatility_multiplier": float(volatility_mult),
        "final_cgri": final,
        "rev_hhi": rev["hhi"], "rev_hhi_sub": rev["hhi_sub"],
        "sc_hhi_sup": sc["hhi_sup"], "sc_hhi_fac": sc["hhi_fac"],
        "sc_hhi_combined": sc["hhi_combined"], "sc_hhi_sub": sc["hhi_sub"],
        "sc_sup_component": sc["sup_component"], "sc_fac_component": sc["fac_component"],
        "sc_intermediate": sc["intermediate"],
        "rev_shares": rev["std_shares"], "sup_shares": sc["sup_std"],
        "fac_sup_shares": sc["fac_std"],
    }


# ═══════════════════════════════════════════════════════════════════════════
# DATA LOADING
# ═══════════════════════════════════════════════════════════════════════════

def _ws(path, sheet):
    return load_workbook(path, data_only=True)[sheet]


@st.cache_data
def load_all():
    ws = _ws(DATA_DIR / "HQ Country Risk Index.xlsx", "Geopolitical Risk Index")
    country_rows = []
    for r in range(2, ws.max_row + 1):
        c, s = ws.cell(r, 1).value, ws.cell(r, 2).value
        if c is not None and isinstance(s, (int, float)):
            country_rows.append({"country": str(c).strip(), "gri": float(s)})
    country_df   = pd.DataFrame(country_rows).drop_duplicates("country").sort_values("country")
    country_risk = dict(zip(country_df["country"], country_df["gri"]))

    ws2 = _ws(DATA_DIR / "Sector Risk Multiplier.xlsx", "Sector Risk Multiplier")

    # Multiplier mapping from S&P Global Industry Risk Assessment label
    _SP_LABEL_TO_MULT = {
        "Very low risk":      0.75,
        "Low risk":           0.85,
        "Intermediate risk":  1.00,
        "Moderately high risk": 1.10,
        "High risk":          1.15,
        "Very high risk":     1.25,
    }

    # Build full 53-sector lookup from the left-hand Sector-List columns (A & B)
    sector_mult: dict[str, float] = {}
    for r in range(5, ws2.max_row + 1):
        sec_name  = ws2.cell(r, 1).value
        sp_label  = ws2.cell(r, 2).value
        if sec_name and sp_label:
            sec_str = str(sec_name).strip()
            lab_str = str(sp_label).strip()
            if lab_str in _SP_LABEL_TO_MULT:
                sector_mult[sec_str] = _SP_LABEL_TO_MULT[lab_str]

    # Also include any benchmark-company sectors not in the S&P 53-sector list
    # (e.g. "Insurance (Financial services)" used by Allianz)
    for r in range(5, ws2.max_row + 1):
        se, la, mu = ws2.cell(r,5).value, ws2.cell(r,6).value, ws2.cell(r,8).value
        if se and isinstance(mu, (int, float)):
            sec_str = str(se).strip()
            if sec_str not in sector_mult:
                sector_mult[sec_str] = float(mu)

    # Build company→sector mapping from the right-hand company columns (D–H)
    sec_rows = []
    for r in range(5, ws2.max_row + 1):
        co, se, la, mu = ws2.cell(r,4).value, ws2.cell(r,5).value, ws2.cell(r,6).value, ws2.cell(r,8).value
        if co and se and isinstance(mu, (int, float)):
            sec_rows.append({"company": str(co).strip(), "sector": str(se).strip(),
                             "sp_label": str(la).strip() if la else "", "multiplier": float(mu)})
    sector_df = pd.DataFrame(sec_rows)

    ws3 = _ws(DATA_DIR / "Template CGRI.xlsx", "Final Template - 2024")
    bench = []
    for r in range(3, ws3.max_row + 1):
        co = ws3.cell(r, 1).value
        if co is None: continue
        try:
            bench.append({
                "Company":               str(co).strip(),
                "HQ Risk":               float(ws3.cell(r, 2).value),
                "Revenue Exposure":      float(ws3.cell(r, 3).value),
                "Supply Chain":          float(ws3.cell(r, 4).value),
                "Financial Multiplier":  float(ws3.cell(r, 5).value),
                "Sector Multiplier":     float(ws3.cell(r, 6).value),
                "Volatility Multiplier": float(ws3.cell(r, 7).value),
                "Final CGRI":            float(ws3.cell(r, 8).value),
            })
        except (TypeError, ValueError):
            continue
    bench_df = pd.DataFrame(bench)
    sec_co_map = dict(zip(sector_df["company"], sector_df["sector"]))
    # Name aliases: Template CGRI uses "Total Energies", other sheets use "Totale Energies"
    sec_co_map["Total Energies"] = sec_co_map.get("Totale Energies", "Oil and gas exploration and production")
    bench_df["Sector"] = bench_df["Company"].map(sec_co_map).fillna("Unknown")
    bench_df["Risk Category"] = pd.Categorical(
        bench_df["Final CGRI"].apply(lambda s: risk_label(s)[0]),
        categories=["Low", "Moderate", "High", "Very High"],
        ordered=True,
    )
    bench_df = bench_df.sort_values("Final CGRI", ascending=False).reset_index(drop=True)

    ws4 = _ws(DATA_DIR / "Template CGRI.xlsx", "Volatility Risk Multiplier")
    vol_mult = None
    for r in range(5, ws4.max_row + 1):
        yr, mu = ws4.cell(r, 4).value, ws4.cell(r, 8).value
        if yr == 2024 and isinstance(mu, (int, float)):
            vol_mult = float(mu)
            break

    return country_df, country_risk, sector_df, sector_mult, bench_df, vol_mult


# ═══════════════════════════════════════════════════════════════════════════
# UI HELPERS
# ═══════════════════════════════════════════════════════════════════════════

RISK_COLORS = {
    "Low":       "#27ae60",
    "Moderate":  "#f39c12",
    "High":      "#e67e22",
    "Very High": "#e74c3c",
}
COMPONENT_KEYS = ["HQ Risk", "Revenue Exposure", "Supply Chain"]


def risk_label(score: float) -> tuple[str, str]:
    if score < 3.5: return "Low",       RISK_COLORS["Low"]
    if score < 5.0: return "Moderate",  RISK_COLORS["Moderate"]
    if score < 6.5: return "High",      RISK_COLORS["High"]
    return "Very High", RISK_COLORS["Very High"]


def kpi_card(label: str, value: str, sub: str = "", accent: str = "#4b6fff"):
    st.markdown(
        f"""<div class="kpi-card" style="border-left-color:{accent}">
        <div class="kpi-label">{label}</div>
        <div class="kpi-value">{value}</div>
        <div class="kpi-sub">{sub}</div>
        </div>""",
        unsafe_allow_html=True,
    )


def comp_card(label: str, value: str, detail: str = ""):
    st.markdown(
        f"""<div class="comp-card">
        <div class="cl">{label}</div>
        <div class="cv">{value}</div>
        <div class="cd">{detail}</div>
        </div>""",
        unsafe_allow_html=True,
    )


def radar_chart(rows: list[dict], title: str = "") -> go.Figure:
    dims = ["HQ Risk", "Revenue Exposure", "Supply Chain"]
    # 25-colour palette — Alphabet (26) covers all benchmark companies
    pal  = px.colors.qualitative.Alphabet
    n    = len(rows)
    many = n > 6   # switch to lines-only when crowded

    fig = go.Figure()
    for i, row in enumerate(rows):
        vals = [row[d] for d in dims]
        col  = pal[i % len(pal)]
        fig.add_trace(go.Scatterpolar(
            r=vals + [vals[0]], theta=dims + [dims[0]],
            fill="toself" if not many else "none",
            name=row.get("Company", f"#{i+1}"),
            line=dict(color=col, width=2 if many else 2.5),
            fillcolor=col,
            opacity=0.15 if not many else 1.0,
        ))

    # Legend: right-side vertical for many companies, horizontal below for few
    if many:
        legend_cfg = dict(orientation="v", yanchor="top", y=1.0,
                          xanchor="left", x=1.02, font=dict(size=11))
        r_margin = 160
    else:
        legend_cfg = dict(orientation="h", yanchor="bottom", y=-0.22,
                          xanchor="center", x=0.5, font=dict(size=12))
        r_margin = 30

    fig.update_layout(
        polar=dict(
            bgcolor="#f7f8fc",
            radialaxis=dict(visible=True, range=[0, 10],
                            tickfont=dict(size=10), gridcolor="#dde1ee"),
            angularaxis=dict(gridcolor="#dde1ee"),
        ),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        showlegend=True,
        legend=legend_cfg,
        title=dict(text=title, font=dict(size=14, color="#1a1d2e"), x=0.5, xanchor="center"),
        height=500 if many else 400,
        margin=dict(l=30, r=r_margin, t=50, b=80 if not many else 20),
    )
    return fig


def gauge_chart(score: float, company: str) -> go.Figure:
    cat, col = risk_label(score)
    fig = go.Figure(go.Indicator(
        mode="gauge+number",
        value=score,
        number={"font": {"size": 52, "color": col}, "suffix": "", "valueformat": ".2f"},
        gauge={
            "axis": {"range": [0, 10], "tickwidth": 1, "tickcolor": "#8a94a6",
                     "tickfont": {"size": 11}},
            "bar": {"color": col, "thickness": 0.28},
            "bgcolor": "#f0f2f8",
            "borderwidth": 0,
            "steps": [
                {"range": [0, 3.5],  "color": "#d6f5e3"},
                {"range": [3.5, 5.0], "color": "#fef3d8"},
                {"range": [5.0, 6.5], "color": "#fde8d5"},
                {"range": [6.5, 10],  "color": "#fad7d5"},
            ],
            "threshold": {"line": {"color": col, "width": 4}, "thickness": 0.82, "value": score},
        },
        title={"text": f"<b>{cat} Risk</b><br><span style='font-size:0.85em;color:#8a94a6'>{company}</span>",
               "font": {"size": 15, "color": "#1a1d2e"}},
    ))
    fig.update_layout(
        paper_bgcolor="rgba(0,0,0,0)",
        height=300,
        margin=dict(l=30, r=30, t=60, b=20),
    )
    return fig


def init_rows(key: str, n: int = 1):
    if key not in st.session_state:
        st.session_state[key] = [{"country": None, "value": 0.0} for _ in range(n)]


def add_row(key: str):
    st.session_state[key].append({"country": None, "value": 0.0})


def delete_row(key: str, idx: int):
    st.session_state[key].pop(idx)


def country_input_table(key: str, country_options: list) -> dict:
    rows   = st.session_state[key]
    result = {}
    total  = sum(r["value"] for r in rows if r["country"] and r["value"] > 0)

    hcols = st.columns([3, 2, 1, 0.4])
    hcols[0].markdown("<div class='section-hdr'>Country</div>", unsafe_allow_html=True)
    hcols[1].markdown("<div class='section-hdr'>Weight</div>", unsafe_allow_html=True)
    hcols[2].markdown("<div class='section-hdr'>Share</div>", unsafe_allow_html=True)

    delete_idx = None
    for i, row in enumerate(rows):
        c1, c2, c3, c4 = st.columns([3, 2, 1, 0.4])
        idx = 0
        if row["country"] and row["country"] in country_options:
            idx = country_options.index(row["country"]) + 1
        country = c1.selectbox("Country", options=[""] + country_options,
                               index=idx, key=f"{key}_c_{i}", label_visibility="collapsed")
        value   = c2.number_input("Weight", min_value=0.0, step=1.0,
                                  value=float(row["value"]), key=f"{key}_v_{i}",
                                  label_visibility="collapsed")
        share_txt = f"{value / total * 100:.1f}%" if total > 0 and country and value > 0 else "—"
        c3.markdown(
            f"<div style='padding-top:9px;color:#8a94a6;font-size:0.85rem;text-align:center'>{share_txt}</div>",
            unsafe_allow_html=True,
        )
        if c4.button("✕", key=f"{key}_del_{i}", help="Remove this row"):
            delete_idx = i

        row["country"] = country or None
        row["value"]   = value
        if country and value > 0:
            result[country] = value

    if delete_idx is not None:
        delete_row(key, delete_idx)
        st.rerun()

    if total > 0 and abs(total - 100) > 0.5:
        st.caption(f"⚠ Weights sum to {total:.1f} — auto-normalised to 100 %.")

    return result


# ═══════════════════════════════════════════════════════════════════════════
# LOAD DATA
# ═══════════════════════════════════════════════════════════════════════════

country_df, country_risk, sector_df, sector_mult, bench_df, vol_mult = load_all()
country_options = sorted(country_risk.keys())
sector_options  = sorted(sector_mult.keys())


# ═══════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ═══════════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown("## 🌐 CGRI Tool")
    st.markdown("<div style='font-size:0.78rem;color:#6b728a;margin-bottom:18px'>Company Geopolitical Risk Index · 2024 Edition</div>", unsafe_allow_html=True)

    page = st.radio(
        "Navigate",
        ["📊 Benchmark Dashboard", "🧮 Custom Calculator", "ℹ Methodology"],
        label_visibility="collapsed",
    )

    st.divider()
    st.markdown(f"<div class='section-hdr' style='color:#6b728a'>Volatility multiplier (2024)</div>", unsafe_allow_html=True)
    st.markdown(f"<div style='font-size:1.1rem;font-weight:700;color:#d0d4e8'>{vol_mult:.4f}</div>", unsafe_allow_html=True)
    st.markdown("<div style='font-size:0.72rem;color:#6b728a'>CBOE VIX via FRED</div>", unsafe_allow_html=True)

    st.divider()

    st.markdown("<div class='section-hdr' style='color:#6b728a'>Group members</div>", unsafe_allow_html=True)
    members = [
        "Calogero Emanuele Ferrante",
        "Arzum Karahan",
        "Andrea Lorusso",
        "Angela Lorusso",
        "Van Anh Nguyen",
        "Valerio Parigi",
        "Domenico Soprano",
        "Pietro Zini",
    ]
    st.markdown(
        "".join(f"<div style='font-size:0.75rem;color:#8a94a6;padding:2px 0'>{m}</div>" for m in members),
        unsafe_allow_html=True,
    )

    st.divider()


# ═══════════════════════════════════════════════════════════════════════════
# PAGE: BENCHMARK DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════

if page == "📊 Benchmark Dashboard":

    st.markdown("## Benchmark Dashboard")
    st.markdown("25-company portfolio · 2024 CGRI scores")

    # ── Summary KPIs ────────────────────────────────────────────────────────
    avg_cgri = bench_df["Final CGRI"].mean()
    max_row  = bench_df.loc[bench_df["Final CGRI"].idxmax()]
    min_row  = bench_df.loc[bench_df["Final CGRI"].idxmin()]
    n_high   = (bench_df["Risk Category"].isin(["High", "Very High"])).sum()

    k1, k2, k3, k4 = st.columns(4)
    with k1: kpi_card("Portfolio Average", f"{avg_cgri:.2f}", "mean CGRI score", "#4b6fff")
    with k2: kpi_card("Highest Risk", f"{max_row['Final CGRI']:.2f}", max_row["Company"], RISK_COLORS["Very High"])
    with k3: kpi_card("Lowest Risk",  f"{min_row['Final CGRI']:.2f}", min_row["Company"],  RISK_COLORS["Low"])
    with k4: kpi_card("High / Very High", str(n_high), f"of {len(bench_df)} companies", RISK_COLORS["High"])

    st.markdown("---")

    # ── Filters ─────────────────────────────────────────────────────────────
    with st.expander("🔍 Filters", expanded=False):
        f1, f2, f3 = st.columns([2, 2, 1])
        risk_filter   = f1.multiselect("Risk category", list(RISK_COLORS.keys()), default=list(RISK_COLORS.keys()))
        sector_filter = f2.multiselect("Sector", sorted(bench_df["Sector"].unique()), default=sorted(bench_df["Sector"].unique()))
        asc           = f3.checkbox("Sort ascending", value=False)

    view = bench_df[
        bench_df["Risk Category"].isin(risk_filter) &
        bench_df["Sector"].isin(sector_filter)
    ].sort_values("Final CGRI", ascending=asc)

    if view.empty:
        st.warning("No companies match the current filters.")
        st.stop()

    # ── Main bar chart ───────────────────────────────────────────────────────
    fig_bar = px.bar(
        view, x="Company", y="Final CGRI",
        color="Risk Category", color_discrete_map=RISK_COLORS,
        hover_data=["Sector"] + COMPONENT_KEYS + ["Sector Multiplier"],
        height=420,
    )
    fig_bar.update_layout(
        legend_title="Risk Category",
        plot_bgcolor="#f7f8fc",
        paper_bgcolor="rgba(0,0,0,0)",
        title=dict(text="Final CGRI Score — Ranked Portfolio", font=dict(size=15, color="#1a1d2e")),
        xaxis=dict(showgrid=False, tickangle=-90, tickfont=dict(size=11), automargin=True),
        yaxis=dict(gridcolor="#e4e7f0", title="CGRI Score"),
        bargap=0.35,
        margin=dict(b=10),
    )
    st.plotly_chart(fig_bar, use_container_width=True)

    # ── Radar ─────────────────────────────────────────────────────────────────
    st.markdown("#### Risk dimension comparison")
    sel_cos = st.multiselect(
        "Select companies to compare (all 25 supported)",
        options=view["Company"].tolist(),
        default=view["Company"].tolist()[:3],
    )
    if sel_cos:
        radar_rows = (
            view[view["Company"].isin(sel_cos)][["Company"] + COMPONENT_KEYS].to_dict("records")
        )
        st.plotly_chart(radar_chart(radar_rows), use_container_width=True)

    # ── Stacked ───────────────────────────────────────────────────────────────
    st.markdown("#### Weighted component breakdown")
    cb = view.copy()
    cb["HQ Risk (w)"]     = 0.15 * cb["HQ Risk"]
    cb["Revenue (w)"]     = 0.45 * cb["Revenue Exposure"]
    cb["Supply Chain (w)"] = 0.40 * cb["Supply Chain"]
    melted = cb[["Company", "HQ Risk (w)", "Revenue (w)", "Supply Chain (w)"]].melt(
        id_vars="Company", var_name="Component", value_name="Score")
    stack_pal = {"HQ Risk (w)": "#4b6fff", "Revenue (w)": "#00c897",
                 "Supply Chain (w)": "#f05545"}
    fig_stack = px.bar(
        melted, x="Company", y="Score", color="Component",
        color_discrete_map=stack_pal,
        barmode="stack", height=500,
    )
    fig_stack.update_layout(
        plot_bgcolor="#f7f8fc",
        paper_bgcolor="rgba(0,0,0,0)",
        xaxis=dict(showgrid=False, tickangle=-90, tickfont=dict(size=11),
                   automargin=True, title=None),
        yaxis=dict(gridcolor="#e4e7f0", title="Weighted Score"),
        bargap=0.35,
        margin=dict(t=20, b=20, r=160),
        legend=dict(orientation="v", yanchor="middle", y=0.5,
                    xanchor="left", x=1.02,
                    title=None, font=dict(size=12)),
    )
    st.plotly_chart(fig_stack, use_container_width=True)

    # ── Data table ───────────────────────────────────────────────────────────
    st.markdown("#### Full data table")
    disp_cols = ["Company", "Sector", "Risk Category", "Final CGRI"] + COMPONENT_KEYS + \
                ["Financial Multiplier", "Sector Multiplier", "Volatility Multiplier"]
    fmt = {
        "Final CGRI": "{:.2f}", "HQ Risk": "{:.2f}", "Revenue Exposure": "{:.2f}",
        "Supply Chain": "{:.2f}", "Financial Multiplier": "{:.1f}",
        "Sector Multiplier": "{:.2f}", "Volatility Multiplier": "{:.4f}",
    }
    # Colour palette for Risk Category cells
    RC_BG   = {"Low": "#d6f5e3", "Moderate": "#fef3d8", "High": "#fde8d5", "Very High": "#fad7d5"}
    RC_TEXT = {"Low": "#1a6b3c", "Moderate": "#7a5000", "High": "#7a3000", "Very High": "#7a1010"}

    def style_risk(val):
        bg   = RC_BG.get(val, "")
        text = RC_TEXT.get(val, "")
        return f"background-color: {bg}; color: {text}; font-weight: 600; border-radius: 4px;"

    tbl = view[disp_cols].reset_index(drop=True)
    tbl.index = tbl.index + 1
    styled = tbl.style.format(fmt).map(style_risk, subset=["Risk Category"])
    st.dataframe(styled, use_container_width=True, height=380)
    csv_bytes = view[disp_cols].to_csv(index=False).encode()
    st.download_button("⬇ Download CSV", data=csv_bytes,
                       file_name="cgri_benchmark_2024.csv", mime="text/csv")


# ═══════════════════════════════════════════════════════════════════════════
# PAGE: CUSTOM CALCULATOR
# ═══════════════════════════════════════════════════════════════════════════

elif page == "🧮 Custom Calculator":

    st.markdown("## Custom CGRI Calculator")

    # ── Step 1: Company profile ──────────────────────────────────────────────
    st.markdown("### Step 1 — Company profile")
    p1, p2, p3, p4 = st.columns([2, 2, 2, 1.5])
    company_name    = p1.text_input("Company name", value="My Company")
    hq_country      = p2.selectbox("HQ country", country_options)
    sector          = p3.selectbox("Sector (S&P Global)", sector_options)
    net_debt_ebitda = p4.number_input(
        "Net Debt / EBITDA", value=1.0, step=0.1, format="%.2f",
        help="Determines Financial Leverage multiplier: <0 → 0.8 · 0–2 → 0.9 · 2–4 → 1.0 · 4–6 → 1.1 · ≥6 → 1.2",
    )
    fin_prev = net_debt_to_financial_multiplier(net_debt_ebitda)
    p4.caption(f"Leverage multiplier: **×{fin_prev}**")

    st.markdown("---")

    # ── Step 2: Exposure tables ──────────────────────────────────────────────
    st.markdown("### Step 2 — Geographic exposure")
    st.info(
        "**Weights are automatically standardised to sum to 100 %.** "
        "Enter any raw unit (% of sales, USD mn, supplier count, number of sites, etc.). "
        "Rows with a blank country or zero weight are ignored.",
        icon="ℹ️",
    )

    init_rows("rev"); init_rows("sup"); init_rows("supfac")

    # ── Two-column exposure inputs ───────────────────────────────────────────
    col_a, col_b = st.columns(2, gap="large")

    with col_a:
        st.markdown("#### 1  Revenue by country")
        st.caption("Enter revenues per country in any unit (e.g. % of sales, USD mn). Auto-normalised to 100 %.")
        rev_input = country_input_table("rev", country_options)
        if st.button("➕ Add country", key="add_rev"):
            add_row("rev"); st.rerun()

        st.markdown("#### 3  Supplier facility distribution")
        st.caption("Enter supplier-side facility count or % per country. Auto-normalised to 100 %.")
        supfac_input = country_input_table("supfac", country_options)
        if st.button("➕ Add country", key="add_supfac"):
            add_row("supfac"); st.rerun()

    with col_b:
        st.markdown("#### 2  Supplier distribution")
        st.caption("Enter supplier count or % per country. Auto-normalised to 100 %.")
        sup_input = country_input_table("sup", country_options)
        if st.button("➕ Add country", key="add_sup"):
            add_row("sup"); st.rerun()

    st.markdown("---")

    # ── Step 3: Compute ──────────────────────────────────────────────────────
    st.markdown("### Step 3 — Compute")
    run_calc = st.button("🔍 Compute CGRI", type="primary")

    if run_calc:
        missing = []
        if not rev_input:    missing.append("Revenue by country")
        if not sup_input:    missing.append("Supplier distribution")
        if not supfac_input: missing.append("Supplier facility distribution")
        if missing:
            st.warning(f"⚠ Please fill in: **{', '.join(missing)}** before computing.")
            st.stop()
        try:
            result = compute_cgri(
                hq_country=hq_country, sector=sector,
                net_debt_ebitda=net_debt_ebitda,
                revenue_by_country=rev_input,
                suppliers_by_country=sup_input,
                supplier_facilities_by_country=supfac_input,
                country_risk=country_risk,
                sector_mult_lookup=sector_mult,
                volatility_mult=vol_mult,
                company_name=company_name,
            )

            final = result["final_cgri"]
            cat, col = risk_label(final)

            st.markdown(f"## Results — {company_name}")

            # ── Gauge + component cards ──────────────────────────────────────
            g_col, m_col = st.columns([1, 1], gap="large")

            with g_col:
                st.plotly_chart(gauge_chart(final, company_name), use_container_width=True)

            with m_col:
                st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)
                r1a, r1b = st.columns(2)
                with r1a:
                    comp_card("HQ Risk", f"{result['hq_risk']:.2f}", hq_country)
                with r1b:
                    comp_card("Revenue Exposure", f"{result['revenue_exposure']:.2f}",
                              f"HHI {result['rev_hhi']:.2f} → ×{result['rev_hhi_sub']:.2f}")
                st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
                r2a, r2b = st.columns(2)
                with r2a:
                    comp_card("Supply Chain", f"{result['supply_chain']:.2f}",
                              f"HHI {result['sc_hhi_combined']:.2f} → ×{result['sc_hhi_sub']:.2f}")
                with r2b:
                    comp_card("Financial Leverage", f"×{result['financial_multiplier']:.1f}",
                              f"Net D/EBITDA {net_debt_ebitda:.2f}")
                st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
                r3a, r3b = st.columns(2)
                with r3a:
                    comp_card("Sector Multiplier", f"×{result['sector_multiplier']:.2f}", sector)
                with r3b:
                    comp_card("Volatility Multiplier", f"×{result['volatility_multiplier']:.4f}",
                              "2024 VIX avg (CBOE / FRED)")

            st.markdown("---")

            # ── Radar + benchmark comparison ─────────────────────────────────
            rdr_col, cmp_col = st.columns(2, gap="large")

            with rdr_col:
                st.markdown("#### Risk profile (spider chart)")
                custom_radar = {
                    "Company":          company_name,
                    "HQ Risk":          result["hq_risk"],
                    "Revenue Exposure": result["revenue_exposure"],
                    "Supply Chain":     result["supply_chain"],
                }
                # overlay benchmark average
                bench_avg = {
                    "Company": "Benchmark avg",
                    "HQ Risk":          bench_df["HQ Risk"].mean(),
                    "Revenue Exposure": bench_df["Revenue Exposure"].mean(),
                    "Supply Chain":     bench_df["Supply Chain"].mean(),
                }
                st.plotly_chart(
                    radar_chart([custom_radar, bench_avg], "vs. benchmark average"),
                    use_container_width=True,
                )

            with cmp_col:
                st.markdown("#### Ranking vs. benchmark portfolio")
                custom_row = pd.DataFrame([{
                    "Company": f"▶ {company_name}", "Final CGRI": final, "Risk Category": cat,
                }])
                compare = pd.concat(
                    [bench_df[["Company", "Final CGRI", "Risk Category"]], custom_row],
                    ignore_index=True,
                ).sort_values("Final CGRI", ascending=False)
                fig_cmp = px.bar(
                    compare, x="Company", y="Final CGRI",
                    color="Risk Category", color_discrete_map=RISK_COLORS,
                    height=400,
                )
                fig_cmp.update_layout(
                    showlegend=False,
                    plot_bgcolor="#f7f8fc",
                    paper_bgcolor="rgba(0,0,0,0)",
                    xaxis=dict(showgrid=False, tickangle=-90, tickfont=dict(size=11), automargin=True),
                    yaxis=dict(gridcolor="#e4e7f0", title="CGRI Score"),
                    bargap=0.3,
                    margin=dict(b=10),
                )
                st.plotly_chart(fig_cmp, use_container_width=True)

            # ── Supply chain detail ──────────────────────────────────────────
            with st.expander("Supply chain detail"):
                st.markdown(
                    f"""
| | Value |
|---|---|
| Suppliers component (Σ GRI × sup_share) | {result['sc_sup_component']:.4f} |
| Supplier-facilities component (Σ GRI × fac_share) | {result['sc_fac_component']:.4f} |
| Intermediate (0.5 × A + 0.5 × B) | {result['sc_intermediate']:.4f} |
| HHI suppliers | {result['sc_hhi_sup']:.4f} |
| HHI supplier facilities | {result['sc_hhi_fac']:.4f} |
| HHI combined (avg) | {result['sc_hhi_combined']:.4f} |
| HHI submultiplier | {result['sc_hhi_sub']:.2f} |
| **Supply Chain Exposure** | **{result['supply_chain']:.4f}** |
"""
                )

            # ── Export ───────────────────────────────────────────────────────
            st.markdown("#### Export results")
            summary = {
                "Company": company_name, "HQ Country": result["hq_country"],
                "Sector": result["sector"], "Net Debt/EBITDA": result["net_debt_ebitda"],
                "HQ Risk": round(result["hq_risk"], 4),
                "Revenue Exposure": round(result["revenue_exposure"], 4),
                "Revenue HHI": round(result["rev_hhi"], 4),
                "Revenue HHI Sub": result["rev_hhi_sub"],
                "Supply Chain Exposure": round(result["supply_chain"], 4),
                "SC Sup Component": round(result["sc_sup_component"], 4),
                "SC Fac Component": round(result["sc_fac_component"], 4),
                "SC Intermediate": round(result["sc_intermediate"], 4),
                "SC HHI Combined": round(result["sc_hhi_combined"], 4),
                "SC HHI Sub": result["sc_hhi_sub"],
                "Financial Leverage Multiplier": result["financial_multiplier"],
                "Sector Multiplier": result["sector_multiplier"],
                "Volatility Multiplier": round(result["volatility_multiplier"], 4),
                "Final CGRI": round(result["final_cgri"], 4),
                "Risk Category": cat,
            }
            e1, e2 = st.columns(2)
            e1.download_button(
                "⬇ CSV (summary)",
                data=pd.DataFrame([summary]).to_csv(index=False).encode(),
                file_name=f"{company_name.replace(' ','_')}_cgri.csv",
                mime="text/csv",
            )
            export_json = {k: v for k, v in result.items() if not isinstance(v, dict)}
            export_json.update({"rev_shares": result["rev_shares"],
                                "sup_shares": result["sup_shares"],
                                "fac_sup_shares": result["fac_sup_shares"]})
            e2.download_button(
                "⬇ JSON (full detail)",
                data=json.dumps(export_json, indent=2, default=str),
                file_name=f"{company_name.replace(' ','_')}_cgri.json",
                mime="application/json",
            )

        except Exception as exc:
            st.error(f"⚠ {exc}")
            st.exception(exc)


# ═══════════════════════════════════════════════════════════════════════════
# PAGE: METHODOLOGY
# ═══════════════════════════════════════════════════════════════════════════

elif page == "ℹ Methodology":

    st.markdown("## Methodology & Data Sources")

    st.markdown(r"""
### CGRI Formula (2024)

$$
\text{CGRI} = \bigl(
  0.15 \cdot \text{HQ Risk}
+ 0.45 \cdot \text{Revenue Exposure}
+ 0.40 \cdot \text{Supply Chain Exposure}
\bigr)
\times \text{Sector Multiplier} \times \text{Volatility Multiplier} \times \text{Financial Leverage Multiplier}
$$
""")

    c1, c2 = st.columns(2)
    with c1:
        st.markdown("""
### Component formulas

| Component | Weight | Formula |
|---|---|---|
| **HQ Risk** | 15 % | Country GRI of the headquarters location |
| **Revenue Exposure** | 45 % | Σ(GRI_c × rev_share_c) × HHI_sub |
| **Supply Chain** | 40 % | (0.5 × C_suppliers + 0.5 × C_sup_facilities) × HHI_sub |
| **Sector Multiplier** | × | S&P Global Industry Risk Assessment |
| **Volatility Multiplier** | × | 1 + Δ% vs. long-run VIX average |
| **Financial Leverage** | × | 0.8 / 0.9 / 1.0 / 1.1 / 1.2 from Net Debt / EBITDA |
""")

    with c2:
        st.markdown("""
### HHI Concentration Submultiplier

| HHI | Submultiplier | Interpretation |
|-----|---------------|----------------|
| < 0.15 | **0.90** | Highly diversified |
| 0.15 – 0.25 | **1.00** | Moderate |
| 0.25 – 0.40 | **1.10** | Somewhat concentrated |
| 0.40 – 0.60 | **1.25** | Concentrated |
| ≥ 0.60 | **1.50** | Highly concentrated |

### Financial Leverage Multiplier

| Net Debt / EBITDA | Multiplier |
|-------------------|-----------|
| < 0 | **0.8** |
| 0 – < 2 | **0.9** |
| 2 – < 4 | **1.0** |
| 4 – < 6 | **1.1** |
| ≥ 6 | **1.2** |
""")

    st.markdown("""
### Risk categories

| Score | Category |
|---|---|
| < 3.5 | 🟢 Low |
| 3.5 – 5.0 | 🟡 Moderate |
| 5.0 – 6.5 | 🟠 High |
| ≥ 6.5 | 🔴 Very High |

### Data sources
- **Country GRI scores** — 147 countries ([geopriskindex.com](https://www.geopriskindex.com))
- **Sector multipliers** — S&P Global Industry Risk Assessment
- **Volatility multiplier** — CBOE VIX annual average (FRED), 2024 = **0.9348**
- **Benchmark scores** — 25 global companies (data collected from Bloomberg)
""")
