from pathlib import Path
import io
import json

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl import load_workbook

st.set_page_config(page_title="CGRI Tool", layout="wide", page_icon="🌐")

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data" / "processed"

# ─────────────────────────── helpers ────────────────────────────────────────

def load_ws(path, sheet_name, data_only=True):
    wb = load_workbook(path, data_only=data_only)
    return wb[sheet_name]


def normalize_weights(exposures: dict) -> dict:
    cleaned = {
        str(k).strip(): float(v)
        for k, v in exposures.items()
        if k is not None and pd.notna(v) and float(v) > 0
    }
    total = sum(cleaned.values())
    if total == 0:
        return {}
    return {k: v / total for k, v in cleaned.items()}


def weighted_country_score(exposures: dict, country_risk_lookup: dict) -> float:
    weights = normalize_weights(exposures)
    missing = sorted([c for c in weights if c not in country_risk_lookup])
    if missing:
        raise KeyError(f"Missing country risk scores for: {missing[:10]}")
    return sum(country_risk_lookup[country] * weight for country, weight in weights.items())


def hhi_value(exposures: dict) -> float:
    weights = normalize_weights(exposures)
    return sum(w ** 2 for w in weights.values())


def hhi_submultiplier(hhi: float) -> float:
    if hhi < 0.15:
        return 0.90
    if hhi < 0.25:
        return 1.00
    if hhi < 0.40:
        return 1.10
    if hhi < 0.60:
        return 1.25
    return 1.50


def financial_exposure_from_net_debt(net_debt_to_ebitda: float) -> int:
    x = float(net_debt_to_ebitda)
    if x < -1:
        return 2
    if x < 0:
        return 3
    if x < 1:
        return 4
    if x < 2:
        return 5
    if x < 3:
        return 6
    if x < 4:
        return 7
    if x < 5:
        return 8
    if x < 6:
        return 9
    return 10


def risk_category(score: float) -> tuple[str, str]:
    """Return (label, colour) for a CGRI score."""
    if score < 3.5:
        return "Low", "#2ecc71"
    if score < 5.0:
        return "Moderate", "#f39c12"
    if score < 6.5:
        return "High", "#e67e22"
    return "Very High", "#e74c3c"


def final_cgri_score(
    hq_risk, revenue_exposure, supply_chain_exposure,
    facility_risk, financial_exposure,
    sector_multiplier, volatility_multiplier,
) -> float:
    base = (
        0.15 * hq_risk
        + 0.25 * revenue_exposure
        + 0.25 * supply_chain_exposure
        + 0.15 * facility_risk
        + 0.20 * financial_exposure
    )
    return base * sector_multiplier * volatility_multiplier


def compute_revenue_exposure(revenue_by_country, country_risk_lookup):
    std = normalize_weights(revenue_by_country)
    intermediate = weighted_country_score(std, country_risk_lookup)
    rev_hhi = hhi_value(std)
    sub = hhi_submultiplier(rev_hhi)
    return {
        "standardized_revenue_shares": std,
        "revenue_intermediate": intermediate,
        "revenue_hhi": rev_hhi,
        "revenue_hhi_submultiplier": sub,
        "revenue_exposure": intermediate * sub,
    }


def compute_supply_chain_exposure(suppliers_by_country, supplier_facilities_by_country, country_risk_lookup):
    sup_std = normalize_weights(suppliers_by_country)
    fac_std = normalize_weights(supplier_facilities_by_country)
    sup_comp = weighted_country_score(sup_std, country_risk_lookup)
    fac_comp = weighted_country_score(fac_std, country_risk_lookup)
    intermediate = 0.5 * sup_comp + 0.5 * fac_comp
    combined_hhi = 0.5 * hhi_value(sup_std) + 0.5 * hhi_value(fac_std)
    sub = hhi_submultiplier(combined_hhi)
    return {
        "standardized_supplier_shares": sup_std,
        "standardized_supplier_facility_shares": fac_std,
        "supplier_distribution_component": sup_comp,
        "supplier_facility_component": fac_comp,
        "supply_intermediate": intermediate,
        "supply_hhi": combined_hhi,
        "supply_hhi_submultiplier": sub,
        "supply_chain_exposure": intermediate * sub,
    }


def compute_facility_risk(facility_sites_by_country, country_risk_lookup):
    std = normalize_weights(facility_sites_by_country)
    component = weighted_country_score(std, country_risk_lookup)
    return {"standardized_facility_shares": std, "facility_risk": component}


def compute_company_cgri(
    company_name, hq_country, sector, net_debt_to_ebitda,
    revenue_by_country, suppliers_by_country,
    supplier_facilities_by_country, facility_sites_by_country,
    country_risk_lookup, sector_multiplier_lookup, volatility_multiplier,
):
    hq_country = str(hq_country).strip()
    sector = str(sector).strip()

    if hq_country not in country_risk_lookup:
        raise KeyError(f"HQ country not found: {hq_country}")
    if sector not in sector_multiplier_lookup:
        raise KeyError(f"Sector not found: {sector}")

    hq_risk = float(country_risk_lookup[hq_country])
    rev = compute_revenue_exposure(revenue_by_country, country_risk_lookup)
    sup = compute_supply_chain_exposure(suppliers_by_country, supplier_facilities_by_country, country_risk_lookup)
    fac = compute_facility_risk(facility_sites_by_country, country_risk_lookup)
    fin_exp = financial_exposure_from_net_debt(net_debt_to_ebitda)
    sec_mult = float(sector_multiplier_lookup[sector])

    final = final_cgri_score(
        hq_risk, rev["revenue_exposure"], sup["supply_chain_exposure"],
        fac["facility_risk"], fin_exp, sec_mult, volatility_multiplier,
    )

    return {
        "company": company_name,
        "hq_country": hq_country,
        "sector": sector,
        "net_debt_to_ebitda": float(net_debt_to_ebitda),
        "hq_risk": hq_risk,
        "revenue_exposure": rev["revenue_exposure"],
        "supply_chain_exposure": sup["supply_chain_exposure"],
        "facility_risk": fac["facility_risk"],
        "financial_exposure": fin_exp,
        "sector_multiplier": sec_mult,
        "volatility_multiplier": float(volatility_multiplier),
        "final_risk_index": final,
        "revenue_hhi": rev["revenue_hhi"],
        "revenue_hhi_submultiplier": rev["revenue_hhi_submultiplier"],
        "supply_hhi": sup["supply_hhi"],
        "supply_hhi_submultiplier": sup["supply_hhi_submultiplier"],
        "standardized_revenue_shares": rev["standardized_revenue_shares"],
        "standardized_supplier_shares": sup["standardized_supplier_shares"],
        "standardized_supplier_facility_shares": sup["standardized_supplier_facility_shares"],
        "standardized_facility_shares": fac["standardized_facility_shares"],
    }


# ─────────────────────────── data loading ───────────────────────────────────

@st.cache_data
def load_reference_data():
    files = {
        "template": DATA_DIR / "Template CGRI.xlsx",
        "hq":       DATA_DIR / "HQ Country Risk Index.xlsx",
        "sector":   DATA_DIR / "Sector Risk Multiplier.xlsx",
    }

    # ── Country risk lookup ──────────────────────────────────────────────────
    hq_geo_ws = load_ws(files["hq"], "Geopolitical Risk Index")
    country_rows = []
    for r in range(2, hq_geo_ws.max_row + 1):
        country = hq_geo_ws.cell(r, 1).value
        score   = hq_geo_ws.cell(r, 2).value
        if country is not None and isinstance(score, (int, float)):
            country_rows.append({"country": str(country).strip(), "country_risk": float(score)})
    country_df = (
        pd.DataFrame(country_rows)
        .drop_duplicates(subset=["country"])
        .sort_values("country")
    )
    country_lookup = dict(zip(country_df["country"], country_df["country_risk"]))

    # ── Sector multiplier lookup ─────────────────────────────────────────────
    sector_ws = load_ws(files["sector"], "Sector Risk Multiplier")
    sector_rows = []
    for r in range(5, sector_ws.max_row + 1):
        company_name  = sector_ws.cell(r, 4).value
        company_sec   = sector_ws.cell(r, 5).value
        sec_label     = sector_ws.cell(r, 6).value
        sec_mult      = sector_ws.cell(r, 8).value
        if company_name and company_sec and isinstance(sec_mult, (int, float)):
            sector_rows.append({
                "company":          str(company_name).strip(),
                "sector":           str(company_sec).strip(),
                "sector_risk_label": str(sec_label).strip() if sec_label else "",
                "sector_multiplier": float(sec_mult),
            })
    company_sector_df = pd.DataFrame(sector_rows)
    sector_multiplier_lookup = dict(
        zip(company_sector_df["sector"], company_sector_df["sector_multiplier"])
    )

    # ── Benchmark table from Final Template ──────────────────────────────────
    tmpl_ws = load_ws(files["template"], "Final Template - 2024")
    bench_rows = []
    for r in range(3, tmpl_ws.max_row + 1):
        company = tmpl_ws.cell(r, 1).value
        if company is None:
            continue
        try:
            bench_rows.append({
                "Company":              str(company).strip(),
                "HQ Risk":              float(tmpl_ws.cell(r, 2).value),
                "Revenue Exposure":     float(tmpl_ws.cell(r, 3).value),
                "Supply Chain":         float(tmpl_ws.cell(r, 4).value),
                "Facility Risk":        float(tmpl_ws.cell(r, 5).value),
                "Financial Exposure":   float(tmpl_ws.cell(r, 6).value),
                "Sector Multiplier":    float(tmpl_ws.cell(r, 7).value),
                "Volatility Multiplier":float(tmpl_ws.cell(r, 8).value),
                "Final CGRI":           float(tmpl_ws.cell(r, 9).value),
            })
        except (TypeError, ValueError):
            continue
    benchmark_df = pd.DataFrame(bench_rows).sort_values("Final CGRI", ascending=False)
    benchmark_df["Risk Category"] = benchmark_df["Final CGRI"].apply(lambda s: risk_category(s)[0])

    # ── Sector labels for benchmark ──────────────────────────────────────────
    sec_map = dict(zip(company_sector_df["company"], company_sector_df["sector"]))
    benchmark_df["Sector"] = benchmark_df["Company"].map(sec_map).fillna("Unknown")

    # ── 2024 volatility multiplier ───────────────────────────────────────────
    vol_ws = load_ws(files["template"], "Volatility Risk Multiplier")
    volatility_multiplier = None
    for r in range(5, vol_ws.max_row + 1):
        year = vol_ws.cell(r, 4).value
        mult = vol_ws.cell(r, 8).value
        if year == 2024 and isinstance(mult, (int, float)):
            volatility_multiplier = float(mult)
            break

    return country_df, country_lookup, company_sector_df, sector_multiplier_lookup, benchmark_df, volatility_multiplier


# ─────────────────────────── session-state row helpers ──────────────────────

def init_rows(name, n=3):
    if name not in st.session_state:
        st.session_state[name] = [{"country": None, "value": 0.0} for _ in range(n)]


def add_row(name):
    st.session_state[name].append({"country": None, "value": 0.0})


def collect_rows(name, country_options):
    rows   = st.session_state[name]
    result = {}
    total  = sum(r["value"] for r in rows if r["country"] and r["value"] > 0)

    for i, row in enumerate(rows):
        col_a, col_b, col_c = st.columns([3, 2, 1])
        country = col_a.selectbox(
            "Country",
            options=[""] + country_options,
            index=0 if not row["country"] else (
                country_options.index(row["country"]) + 1
                if row["country"] in country_options else 0
            ),
            key=f"{name}_country_{i}",
            label_visibility="collapsed",
        )
        value = col_b.number_input(
            "Weight (%)",
            min_value=0.0, step=1.0,
            value=float(row["value"]),
            key=f"{name}_value_{i}",
            label_visibility="collapsed",
        )
        share_txt = ""
        if total > 0 and country and value > 0:
            share_txt = f"{value / total * 100:.1f}%"
        col_c.markdown(f"<div style='padding-top:8px;color:#888'>{share_txt}</div>", unsafe_allow_html=True)

        row["country"] = country if country != "" else None
        row["value"] = value
        if country and value > 0:
            result[country] = value

    if total > 0 and abs(total - 100) > 0.5:
        st.caption(f"⚠ Weights sum to {total:.1f} — will be auto-normalised to 100%.")
    return result


# ─────────────────────────── radar chart ────────────────────────────────────

RADAR_DIMS  = ["HQ Risk", "Revenue Exposure", "Supply Chain", "Facility Risk", "Financial Exposure"]
RADAR_KEYS  = ["HQ Risk", "Revenue Exposure", "Supply Chain", "Facility Risk", "Financial Exposure"]


def radar_fig(companies_data: list[dict], title=""):
    """
    companies_data: list of dicts with keys matching RADAR_KEYS plus 'Company'.
    """
    fig = go.Figure()
    colours = px.colors.qualitative.Plotly
    for i, row in enumerate(companies_data):
        vals = [row[k] for k in RADAR_KEYS]
        fig.add_trace(go.Scatterpolar(
            r=vals + [vals[0]],
            theta=RADAR_DIMS + [RADAR_DIMS[0]],
            fill="toself",
            name=row.get("Company", f"Company {i+1}"),
            line_color=colours[i % len(colours)],
            opacity=0.75,
        ))
    fig.update_layout(
        polar=dict(radialaxis=dict(visible=True, range=[0, 10])),
        showlegend=True,
        title=title,
        height=420,
        margin=dict(l=20, r=20, t=50, b=20),
    )
    return fig


# ─────────────────────────── main app ───────────────────────────────────────

(
    country_df, country_lookup,
    company_sector_df, sector_multiplier_lookup,
    benchmark_df, volatility_multiplier,
) = load_reference_data()

country_options = sorted(country_lookup.keys())
sector_options  = sorted(company_sector_df["sector"].unique())

# ── page header ─────────────────────────────────────────────────────────────
st.title("🌐 Corporate Geopolitical Risk Index")
st.caption(
    "Interactive tool — methodology based on the CGRI Excel framework (2024). "
    f"Volatility multiplier (2024): **{volatility_multiplier:.4f}**."
)

tab1, tab2 = st.tabs(["📊 Benchmark Dashboard", "🧮 Custom Calculator"])

# ════════════════════════════════════════════════════════════════════════════
# TAB 1 — DASHBOARD
# ════════════════════════════════════════════════════════════════════════════
with tab1:
    st.subheader("25-Company Benchmark Portfolio")

    # ── filters ─────────────────────────────────────────────────────────────
    col_f1, col_f2, col_f3 = st.columns([2, 2, 1])
    with col_f1:
        risk_filter = st.multiselect(
            "Filter by risk category",
            options=["Low", "Moderate", "High", "Very High"],
            default=["Low", "Moderate", "High", "Very High"],
        )
    with col_f2:
        sector_filter = st.multiselect(
            "Filter by sector",
            options=sorted(benchmark_df["Sector"].unique()),
            default=sorted(benchmark_df["Sector"].unique()),
        )
    with col_f3:
        sort_asc = st.checkbox("Sort ascending", value=False)

    filtered_df = benchmark_df[
        benchmark_df["Risk Category"].isin(risk_filter) &
        benchmark_df["Sector"].isin(sector_filter)
    ].sort_values("Final CGRI", ascending=sort_asc)

    if filtered_df.empty:
        st.warning("No companies match the selected filters.")
    else:
        # ── colour map ───────────────────────────────────────────────────────
        color_map = {"Low": "#2ecc71", "Moderate": "#f39c12", "High": "#e67e22", "Very High": "#e74c3c"}

        # ── bar chart ────────────────────────────────────────────────────────
        fig_bar = px.bar(
            filtered_df,
            x="Company", y="Final CGRI",
            color="Risk Category",
            color_discrete_map=color_map,
            hover_data=["Sector", "HQ Risk", "Revenue Exposure", "Supply Chain",
                        "Facility Risk", "Financial Exposure",
                        "Sector Multiplier", "Volatility Multiplier"],
            title="Final CGRI Score by Company",
            height=420,
        )
        fig_bar.update_layout(xaxis_tickangle=-40, legend_title="Risk Category")
        st.plotly_chart(fig_bar, use_container_width=True)

        # ── component stacked bar ────────────────────────────────────────────
        with st.expander("Component breakdown (weighted)"):
            comp_df = filtered_df.copy()
            comp_df["HQ Risk (w)"]         = 0.15 * comp_df["HQ Risk"]
            comp_df["Revenue (w)"]          = 0.25 * comp_df["Revenue Exposure"]
            comp_df["Supply Chain (w)"]     = 0.25 * comp_df["Supply Chain"]
            comp_df["Facility (w)"]         = 0.15 * comp_df["Facility Risk"]
            comp_df["Financial (w)"]        = 0.20 * comp_df["Financial Exposure"]
            melt_cols = ["Company", "HQ Risk (w)", "Revenue (w)", "Supply Chain (w)", "Facility (w)", "Financial (w)"]
            melted = comp_df[melt_cols].melt(id_vars="Company", var_name="Component", value_name="Weighted Score")
            fig_stack = px.bar(
                melted, x="Company", y="Weighted Score", color="Component",
                title="Weighted component breakdown (before multipliers)",
                height=420,
            )
            fig_stack.update_layout(xaxis_tickangle=-40, barmode="stack")
            st.plotly_chart(fig_stack, use_container_width=True)

        # ── radar comparison ─────────────────────────────────────────────────
        st.markdown("#### Radar comparison")
        radar_companies = st.multiselect(
            "Select up to 5 companies to compare",
            options=filtered_df["Company"].tolist(),
            default=filtered_df["Company"].tolist()[:3],
        )
        if radar_companies:
            sel = filtered_df[filtered_df["Company"].isin(radar_companies)]
            radar_data = sel.rename(columns={"Supply Chain": "Supply Chain"})[
                ["Company"] + RADAR_KEYS
            ].to_dict("records")
            st.plotly_chart(radar_fig(radar_data, title="Risk dimension comparison"), use_container_width=True)

        # ── data table ───────────────────────────────────────────────────────
        st.markdown("#### Full data table")
        display_cols = ["Company", "Sector", "Risk Category", "Final CGRI",
                        "HQ Risk", "Revenue Exposure", "Supply Chain",
                        "Facility Risk", "Financial Exposure",
                        "Sector Multiplier", "Volatility Multiplier"]
        st.dataframe(
            filtered_df[display_cols].reset_index(drop=True).style.format({
                "Final CGRI": "{:.2f}", "HQ Risk": "{:.2f}",
                "Revenue Exposure": "{:.2f}", "Supply Chain": "{:.2f}",
                "Facility Risk": "{:.2f}", "Financial Exposure": "{:.0f}",
                "Sector Multiplier": "{:.2f}", "Volatility Multiplier": "{:.4f}",
            }).applymap(
                lambda v: f"background-color: {color_map.get(v, 'white')};",
                subset=["Risk Category"]
            ),
            use_container_width=True,
        )

        # ── CSV download ─────────────────────────────────────────────────────
        csv_bytes = filtered_df[display_cols].to_csv(index=False).encode()
        st.download_button(
            "⬇ Download table as CSV",
            data=csv_bytes,
            file_name="cgri_benchmark.csv",
            mime="text/csv",
        )

# ════════════════════════════════════════════════════════════════════════════
# TAB 2 — CUSTOM CALCULATOR
# ════════════════════════════════════════════════════════════════════════════
with tab2:
    st.subheader("Build a custom company profile")
    st.caption(
        "Enter country-level exposure weights (any unit — they are auto-normalised to 100%). "
        "Select the S&P sector category that best fits your company."
    )

    # ── company metadata ─────────────────────────────────────────────────────
    c1, c2 = st.columns(2)
    company_name      = c1.text_input("Company name", value="My Company")
    hq_country        = c2.selectbox("HQ country", country_options)
    sector            = c1.selectbox("Sector (S&P Global category)", sector_options)
    net_debt_to_ebitda = c2.number_input("Net debt / EBITDA (2024)", value=1.0, step=0.1, format="%.2f")

    st.divider()

    # ── exposure inputs ──────────────────────��───────────────────────────────
    init_rows("revenue_rows", 3)
    init_rows("supplier_rows", 3)
    init_rows("supplier_facility_rows", 3)
    init_rows("facility_rows", 3)

    exp1, exp2 = st.columns(2)

    with exp1:
        st.markdown("#### Revenue exposure by country")
        st.caption("Enter % of revenues generated in each country")
        col_h1, col_h2, col_h3 = st.columns([3, 2, 1])
        col_h1.markdown("**Country**")
        col_h2.markdown("**Weight**")
        col_h3.markdown("**Share**")
        revenue_by_country = collect_rows("revenue_rows", country_options)
        if st.button("➕ Add revenue country"):
            add_row("revenue_rows"); st.rerun()

    with exp2:
        st.markdown("#### Facility sites by country")
        st.caption("Enter % of owned/operated facility sites in each country")
        col_h1, col_h2, col_h3 = st.columns([3, 2, 1])
        col_h1.markdown("**Country**")
        col_h2.markdown("**Weight**")
        col_h3.markdown("**Share**")
        facility_sites_by_country = collect_rows("facility_rows", country_options)
        if st.button("➕ Add facility country"):
            add_row("facility_rows"); st.rerun()

    st.divider()
    sup1, sup2 = st.columns(2)

    with sup1:
        st.markdown("#### Supplier distribution by country")
        st.caption("Enter % of suppliers located in each country")
        col_h1, col_h2, col_h3 = st.columns([3, 2, 1])
        col_h1.markdown("**Country**")
        col_h2.markdown("**Weight**")
        col_h3.markdown("**Share**")
        suppliers_by_country = collect_rows("supplier_rows", country_options)
        if st.button("➕ Add supplier country"):
            add_row("supplier_rows"); st.rerun()

    with sup2:
        st.markdown("#### Supplier facility distribution by country")
        st.caption("Enter % of supplier facilities located in each country")
        col_h1, col_h2, col_h3 = st.columns([3, 2, 1])
        col_h1.markdown("**Country**")
        col_h2.markdown("**Weight**")
        col_h3.markdown("**Share**")
        supplier_facilities_by_country = collect_rows("supplier_facility_rows", country_options)
        if st.button("➕ Add supplier facility country"):
            add_row("supplier_facility_rows"); st.rerun()

    st.divider()

    # ── compute button ───────────────────────────────────────────────────────
    if st.button("🔍 Compute CGRI", type="primary", use_container_width=True):
        try:
            result = compute_company_cgri(
                company_name=company_name,
                hq_country=hq_country,
                sector=sector,
                net_debt_to_ebitda=net_debt_to_ebitda,
                revenue_by_country=revenue_by_country,
                suppliers_by_country=suppliers_by_country,
                supplier_facilities_by_country=supplier_facilities_by_country,
                facility_sites_by_country=facility_sites_by_country,
                country_risk_lookup=country_lookup,
                sector_multiplier_lookup=sector_multiplier_lookup,
                volatility_multiplier=volatility_multiplier,
            )

            final      = result["final_risk_index"]
            cat_label, cat_colour = risk_category(final)

            st.markdown("---")
            st.markdown(f"### Results for **{company_name}**")

            # ── headline metric ──────────────────────────────────────────────
            metric_col, _, _ = st.columns([1, 1, 1])
            metric_col.markdown(
                f"""
                <div style='background:{cat_colour};padding:20px;border-radius:10px;text-align:center'>
                    <div style='font-size:2.5rem;font-weight:700;color:white'>{final:.2f}</div>
                    <div style='font-size:1.1rem;color:white'>Final CGRI — {cat_label} Risk</div>
                </div>
                """,
                unsafe_allow_html=True,
            )

            st.markdown("#### Component breakdown")
            m1, m2, m3, m4, m5, m6 = st.columns(6)
            m1.metric("HQ Risk",              f"{result['hq_risk']:.2f}")
            m2.metric("Revenue Exposure",     f"{result['revenue_exposure']:.2f}",
                      delta=f"HHI {result['revenue_hhi']:.2f} × {result['revenue_hhi_submultiplier']:.2f}",
                      delta_color="off")
            m3.metric("Supply Chain",         f"{result['supply_chain_exposure']:.2f}",
                      delta=f"HHI {result['supply_hhi']:.2f} × {result['supply_hhi_submultiplier']:.2f}",
                      delta_color="off")
            m4.metric("Facility Risk",        f"{result['facility_risk']:.2f}")
            m5.metric("Financial Exposure",   f"{result['financial_exposure']:.0f}")
            m6.metric("Sector × Vol. mult",   f"{result['sector_multiplier']:.2f} × {result['volatility_multiplier']:.4f}")

            # ── radar + benchmark comparison ─────────────────────────────────
            chart_col, bench_col = st.columns([1, 1])

            with chart_col:
                custom_radar_data = {
                    "Company":          company_name,
                    "HQ Risk":          result["hq_risk"],
                    "Revenue Exposure": result["revenue_exposure"],
                    "Supply Chain":     result["supply_chain_exposure"],
                    "Facility Risk":    result["facility_risk"],
                    "Financial Exposure": result["financial_exposure"],
                }
                st.plotly_chart(
                    radar_fig([custom_radar_data], title=f"{company_name} — Risk Radar"),
                    use_container_width=True,
                )

            with bench_col:
                st.markdown("**Where does this company stand in the benchmark?**")
                bench_compare = benchmark_df[["Company", "Final CGRI", "Risk Category"]].copy()
                custom_row = pd.DataFrame([{
                    "Company": f"▶ {company_name}",
                    "Final CGRI": final,
                    "Risk Category": cat_label,
                }])
                bench_compare = pd.concat([bench_compare, custom_row]).sort_values("Final CGRI", ascending=False)
                bench_compare["Highlight"] = bench_compare["Company"].str.startswith("▶")
                color_map2 = {"Low": "#2ecc71", "Moderate": "#f39c12", "High": "#e67e22", "Very High": "#e74c3c"}
                fig_bench = px.bar(
                    bench_compare,
                    x="Company", y="Final CGRI",
                    color="Risk Category",
                    color_discrete_map=color_map2,
                    title="Custom company vs benchmark",
                    height=380,
                )
                fig_bench.update_layout(xaxis_tickangle=-40, showlegend=False)
                # Outline the custom company bar
                custom_idx = bench_compare[bench_compare["Highlight"]].index
                fig_bench.update_traces(marker_line_width=0)
                st.plotly_chart(fig_bench, use_container_width=True)

            # ── exports ──────────────────────────────────────────────────────
            st.markdown("#### Export results")
            dl1, dl2 = st.columns(2)

            # JSON export (full detail)
            exportable = {k: v for k, v in result.items() if not isinstance(v, dict)}
            exportable_with_shares = result.copy()
            for k in ["standardized_revenue_shares", "standardized_supplier_shares",
                      "standardized_supplier_facility_shares", "standardized_facility_shares"]:
                exportable_with_shares[k] = result.get(k, {})

            dl1.download_button(
                "⬇ Download JSON (full detail)",
                data=json.dumps(exportable_with_shares, indent=2, default=str),
                file_name=f"{company_name.replace(' ', '_')}_cgri.json",
                mime="application/json",
            )

            # CSV export (summary row)
            summary_row = {
                "Company":             company_name,
                "HQ Country":          result["hq_country"],
                "Sector":              result["sector"],
                "Net Debt/EBITDA":     result["net_debt_to_ebitda"],
                "HQ Risk":             round(result["hq_risk"], 4),
                "Revenue Exposure":    round(result["revenue_exposure"], 4),
                "Revenue HHI":         round(result["revenue_hhi"], 4),
                "Revenue HHI Sub":     result["revenue_hhi_submultiplier"],
                "Supply Chain Exp.":   round(result["supply_chain_exposure"], 4),
                "Supply HHI":          round(result["supply_hhi"], 4),
                "Supply HHI Sub":      result["supply_hhi_submultiplier"],
                "Facility Risk":       round(result["facility_risk"], 4),
                "Financial Exposure":  result["financial_exposure"],
                "Sector Multiplier":   result["sector_multiplier"],
                "Volatility Mult.":    round(result["volatility_multiplier"], 4),
                "Final CGRI":          round(result["final_risk_index"], 4),
                "Risk Category":       cat_label,
            }
            csv_buf = io.StringIO()
            pd.DataFrame([summary_row]).to_csv(csv_buf, index=False)
            dl2.download_button(
                "⬇ Download CSV (summary)",
                data=csv_buf.getvalue().encode(),
                file_name=f"{company_name.replace(' ', '_')}_cgri.csv",
                mime="text/csv",
            )

        except Exception as exc:
            st.error(f"⚠ Computation error: {exc}")
            st.exception(exc)

# ── methodology note ─────────────────────────────────────────────────────────
with st.expander("ℹ Methodology notes"):
    st.markdown("""
**Final CGRI formula:**
$$
\\text{CGRI} = \\bigl(0.15 \\cdot \\text{HQ Risk}
              + 0.25 \\cdot \\text{Revenue Exp.}
              + 0.25 \\cdot \\text{Supply Chain Exp.}
              + 0.15 \\cdot \\text{Facility Risk}
              + 0.20 \\cdot \\text{Financial Exp.}\\bigr)
              \\times \\text{Sector Mult.} \\times \\text{Volatility Mult.}
$$

**HHI concentration submultiplier:**

| HHI range | Submultiplier |
|-----------|--------------|
| < 0.15    | 0.90 (diversified) |
| 0.15 – 0.25 | 1.00 |
| 0.25 – 0.40 | 1.10 |
| 0.40 – 0.60 | 1.25 |
| ≥ 0.60    | 1.50 (concentrated) |

**Risk categories:**

| Score | Category |
|-------|----------|
| < 3.5 | 🟢 Low |
| 3.5 – 5.0 | 🟡 Moderate |
| 5.0 – 6.5 | 🟠 High |
| ≥ 6.5 | 🔴 Very High |

**Financial Exposure** is mapped from Net Debt / EBITDA:
scores range from 2 (strongly net cash) to 10 (highly leveraged, x ≥ 6).

**Data sources:** Country GRI 2024, HQ Country Risk Index, Revenue/Supply Chain/Facility Exposure Excel data,
Sector Risk based on S&P Global Industry Risk Assessment, Volatility from CBOE VIX (FRED: VIXCLS).
    """)
