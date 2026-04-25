"""
Corporate Geopolitical Risk Index (CGRI) — Interactive Tool
============================================================
Methodology strictly follows the Excel template in data/processed/.

Formula
-------
CGRI = (0.15 × HQ_Risk
       + 0.25 × Revenue_Exposure
       + 0.25 × SupplyChain_Exposure
       + 0.15 × Facility_Risk
       + 0.20 × Financial_Exposure)
       × Sector_Multiplier
       × Volatility_Multiplier

Component formulas
------------------
HQ Risk             : country GRI of headquarters country
Revenue Exposure    : Σ(GRI_c × rev_std_c) × HHI_sub(Σ rev_std_c²)
Supply Chain        : 0.5 × sup_comp × 0.5 × fac_comp × HHI_sub(avg_HHI)
                      where avg_HHI = (HHI_suppliers + HHI_facilities) / 2
                      NOTE: Excel formula is 0.5*A*0.5*B = 0.25*A*B (not 0.5*A+0.5*B)
Facility Risk       : Σ(GRI_c × fac_std_c)  [no HHI adjustment]
Financial Exposure  : discrete score 2–10 derived from Net Debt / EBITDA
"""

from pathlib import Path
import io
import json

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl import load_workbook

st.set_page_config(page_title="CGRI Tool", layout="wide", page_icon="🌐")

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data" / "processed"

# ═══════════════════════════════════════════════════════════════════════════
# CORE COMPUTATION FUNCTIONS  (mirror the Excel formulas exactly)
# ═══════════════════════════════════════════════════════════════════════════

def normalize_weights(exposures: dict) -> dict:
    """Drop zeros/nulls, divide each by total so weights sum to 1."""
    cleaned = {
        str(k).strip(): float(v)
        for k, v in exposures.items()
        if k is not None and pd.notna(v) and float(v) > 0
    }
    total = sum(cleaned.values())
    return {k: v / total for k, v in cleaned.items()} if total > 0 else {}


def weighted_country_gri(shares: dict, country_risk: dict) -> float:
    """Return Σ(GRI_c × share_c) — shares must already sum to 1."""
    missing = [c for c in shares if c not in country_risk]
    if missing:
        raise KeyError(f"Countries not in GRI dataset: {missing[:5]}")
    return sum(country_risk[c] * w for c, w in shares.items())


def hhi(shares: dict) -> float:
    """Herfindahl-Hirschman Index = Σ(share_c²)."""
    return sum(w ** 2 for w in shares.values())


def hhi_submultiplier(h: float) -> float:
    """
    Map HHI value to concentration submultiplier (from Excel HHI table).
    < 0.15  → 0.90  (diversified, reduces risk)
    < 0.25  → 1.00
    < 0.40  → 1.10
    < 0.60  → 1.25
    ≥ 0.60  → 1.50  (concentrated, amplifies risk)
    """
    if h < 0.15: return 0.90
    if h < 0.25: return 1.00
    if h < 0.40: return 1.10
    if h < 0.60: return 1.25
    return 1.50


def net_debt_to_financial_score(ratio: float) -> int:
    """
    Map Net Debt / EBITDA to discrete Financial Exposure score (2–10).
    Matches the Excel Financial Exposure mapping table.
    """
    x = float(ratio)
    if x < -1: return 2
    if x <  0: return 3
    if x <  1: return 4
    if x <  2: return 5
    if x <  3: return 6
    if x <  4: return 7
    if x <  5: return 8
    if x <  6: return 9
    return 10


def compute_revenue_exposure(revenue_by_country: dict, country_risk: dict) -> dict:
    """
    Revenue Exposure = Σ(GRI_c × rev_std_c) × HHI_sub(Σ rev_std_c²)
    """
    std = normalize_weights(revenue_by_country)
    intermediate = weighted_country_gri(std, country_risk)
    rev_hhi = hhi(std)
    sub = hhi_submultiplier(rev_hhi)
    return {
        "std_shares": std,
        "intermediate": intermediate,
        "hhi": rev_hhi,
        "hhi_sub": sub,
        "final": intermediate * sub,
    }


def compute_supply_chain_exposure(
    suppliers_by_country: dict,
    supplier_facilities_by_country: dict,
    country_risk: dict,
) -> dict:
    """
    Supply Chain Exposure replicates the Excel formula exactly:

      Component_suppliers  = Σ(GRI_c × sup_std_c)
      Component_facilities = Σ(GRI_c × fac_std_c)
      Intermediate         = 0.5 × C_sup × 0.5 × C_fac   ← Excel formula: 0.5*A*0.5*B
      HHI_combined         = (HHI_sup + HHI_fac) / 2
      Final                = Intermediate × HHI_sub(HHI_combined)

    ⚠ The Excel formula for Intermediate uses multiplication, not addition:
      Excel cell = 0.5*A*0.5*B = 0.25*A*B
      The mathematically expected average would be 0.5*A + 0.5*B.
      We replicate the Excel formula to stay consistent with the benchmark data.
    """
    sup_std = normalize_weights(suppliers_by_country)
    fac_std = normalize_weights(supplier_facilities_by_country)

    sup_comp = weighted_country_gri(sup_std, country_risk)
    fac_comp = weighted_country_gri(fac_std, country_risk)

    # Excel formula: 0.5 × A × 0.5 × B (product, not sum)
    intermediate = 0.5 * sup_comp * 0.5 * fac_comp

    hhi_sup = hhi(sup_std)
    hhi_fac = hhi(fac_std)
    hhi_combined = (hhi_sup + hhi_fac) / 2
    sub = hhi_submultiplier(hhi_combined)

    return {
        "sup_std": sup_std,
        "fac_std": fac_std,
        "sup_component": sup_comp,
        "fac_component": fac_comp,
        "intermediate": intermediate,
        "hhi_sup": hhi_sup,
        "hhi_fac": hhi_fac,
        "hhi_combined": hhi_combined,
        "hhi_sub": sub,
        "final": intermediate * sub,
    }


def compute_facility_risk(facility_by_country: dict, country_risk: dict) -> dict:
    """
    Facility Risk = Σ(GRI_c × fac_std_c)
    No HHI adjustment — matches the Excel Facility Sites Location Risk sheet.
    """
    std = normalize_weights(facility_by_country)
    component = weighted_country_gri(std, country_risk)
    return {"std_shares": std, "final": component}


def compute_cgri(
    hq_country: str,
    sector: str,
    net_debt_ebitda: float,
    revenue_by_country: dict,
    suppliers_by_country: dict,
    supplier_facilities_by_country: dict,
    facility_by_country: dict,
    country_risk: dict,
    sector_mult_lookup: dict,
    volatility_mult: float,
    company_name: str = "Custom Corporate",
) -> dict:
    hq_country = str(hq_country).strip()
    sector = str(sector).strip()

    if hq_country not in country_risk:
        raise KeyError(f"HQ country '{hq_country}' not found in GRI dataset.")
    if sector not in sector_mult_lookup:
        raise KeyError(f"Sector '{sector}' not found in sector multiplier table.")

    hq_risk = float(country_risk[hq_country])
    rev     = compute_revenue_exposure(revenue_by_country, country_risk)
    sc      = compute_supply_chain_exposure(suppliers_by_country, supplier_facilities_by_country, country_risk)
    fac     = compute_facility_risk(facility_by_country, country_risk)
    fin     = net_debt_to_financial_score(net_debt_ebitda)
    sec_m   = float(sector_mult_lookup[sector])

    base = (
        0.15 * hq_risk
        + 0.25 * rev["final"]
        + 0.25 * sc["final"]
        + 0.15 * fac["final"]
        + 0.20 * fin
    )
    final = base * sec_m * volatility_mult

    return {
        "company":              company_name,
        "hq_country":           hq_country,
        "sector":               sector,
        "net_debt_ebitda":      float(net_debt_ebitda),
        # component values
        "hq_risk":              hq_risk,
        "revenue_exposure":     rev["final"],
        "supply_chain":         sc["final"],
        "facility_risk":        fac["final"],
        "financial_exposure":   fin,
        # multipliers
        "sector_multiplier":    sec_m,
        "volatility_multiplier":float(volatility_mult),
        # final
        "final_cgri":           final,
        # HHI details
        "rev_hhi":              rev["hhi"],
        "rev_hhi_sub":          rev["hhi_sub"],
        "sc_hhi_sup":           sc["hhi_sup"],
        "sc_hhi_fac":           sc["hhi_fac"],
        "sc_hhi_combined":      sc["hhi_combined"],
        "sc_hhi_sub":           sc["hhi_sub"],
        "sc_sup_component":     sc["sup_component"],
        "sc_fac_component":     sc["fac_component"],
        "sc_intermediate":      sc["intermediate"],
        # standardized shares (for display)
        "rev_shares":           rev["std_shares"],
        "sup_shares":           sc["sup_std"],
        "fac_sup_shares":       sc["fac_std"],
        "fac_shares":           fac["std_shares"],
    }


# ═══════════════════════════════════════════════════════════════════════════
# DATA LOADING
# ═══════════════════════════════════════════════════════════════════════════

def _ws(path, sheet):
    return load_workbook(path, data_only=True)[sheet]


@st.cache_data
def load_all():
    # ── Country GRI (147 countries) ─────────────────────────────────────────
    ws = _ws(DATA_DIR / "HQ Country Risk Index.xlsx", "Geopolitical Risk Index")
    country_rows = []
    for r in range(2, ws.max_row + 1):
        c, s = ws.cell(r, 1).value, ws.cell(r, 2).value
        if c is not None and isinstance(s, (int, float)):
            country_rows.append({"country": str(c).strip(), "gri": float(s)})
    country_df     = pd.DataFrame(country_rows).drop_duplicates("country").sort_values("country")
    country_risk   = dict(zip(country_df["country"], country_df["gri"]))

    # ── Sector multipliers ──────────────────────────────────────────────────
    ws2 = _ws(DATA_DIR / "Sector Risk Multiplier.xlsx", "Sector Risk Multiplier")
    sec_rows = []
    for r in range(5, ws2.max_row + 1):
        co = ws2.cell(r, 4).value
        se = ws2.cell(r, 5).value
        la = ws2.cell(r, 6).value
        mu = ws2.cell(r, 8).value
        if co and se and isinstance(mu, (int, float)):
            sec_rows.append({
                "company":    str(co).strip(),
                "sector":     str(se).strip(),
                "sp_label":   str(la).strip() if la else "",
                "multiplier": float(mu),
            })
    sector_df   = pd.DataFrame(sec_rows)
    sector_mult = dict(zip(sector_df["sector"], sector_df["multiplier"]))

    # ── Benchmark table (Final Template - 2024) ─────────────────────────────
    ws3 = _ws(DATA_DIR / "Template CGRI.xlsx", "Final Template - 2024")
    bench = []
    for r in range(3, ws3.max_row + 1):
        co = ws3.cell(r, 1).value
        if co is None: continue
        try:
            bench.append({
                "Company":              str(co).strip(),
                "HQ Risk":              float(ws3.cell(r, 2).value),
                "Revenue Exposure":     float(ws3.cell(r, 3).value),
                "Supply Chain":         float(ws3.cell(r, 4).value),
                "Facility Risk":        float(ws3.cell(r, 5).value),
                "Financial Exposure":   float(ws3.cell(r, 6).value),
                "Sector Multiplier":    float(ws3.cell(r, 7).value),
                "Volatility Multiplier":float(ws3.cell(r, 8).value),
                "Final CGRI":           float(ws3.cell(r, 9).value),
            })
        except (TypeError, ValueError):
            continue
    bench_df = pd.DataFrame(bench)

    # Add sector and risk category labels
    sec_co_map = dict(zip(sector_df["company"], sector_df["sector"]))
    bench_df["Sector"] = bench_df["Company"].map(sec_co_map).fillna("Unknown")
    bench_df["Risk Category"] = bench_df["Final CGRI"].apply(lambda s: risk_label(s)[0])
    bench_df = bench_df.sort_values("Final CGRI", ascending=False).reset_index(drop=True)

    # ── 2024 Volatility multiplier ──────────────────────────────────────────
    ws4 = _ws(DATA_DIR / "Template CGRI.xlsx", "Volatility Risk Multiplier")
    vol_mult = None
    for r in range(5, ws4.max_row + 1):
        yr, mu = ws4.cell(r, 4).value, ws4.cell(r, 8).value
        if yr == 2024 and isinstance(mu, (int, float)):
            vol_mult = float(mu)
            break

    return country_df, country_risk, sector_df, sector_mult, bench_df, vol_mult


# ═══════════════════════════════════════════════════════════════════════════
# UI HELPERS
# ═══════════════════════════════════════════════════════════════════════════

RISK_COLORS = {
    "Low":       "#27ae60",
    "Moderate":  "#f39c12",
    "High":      "#e67e22",
    "Very High": "#e74c3c",
}

COMPONENT_KEYS = [
    "HQ Risk", "Revenue Exposure", "Supply Chain",
    "Facility Risk", "Financial Exposure",
]


def risk_label(score: float) -> tuple[str, str]:
    if score < 3.5: return "Low",       RISK_COLORS["Low"]
    if score < 5.0: return "Moderate",  RISK_COLORS["Moderate"]
    if score < 6.5: return "High",      RISK_COLORS["High"]
    return "Very High", RISK_COLORS["Very High"]


def radar_chart(rows: list[dict], title: str = "") -> go.Figure:
    dims = ["HQ Risk", "Revenue Exposure", "Supply Chain", "Facility Risk", "Financial Exposure"]
    fig  = go.Figure()
    pal  = px.colors.qualitative.Plotly
    for i, row in enumerate(rows):
        vals = [row[d] for d in dims]
        fig.add_trace(go.Scatterpolar(
            r=vals + [vals[0]],
            theta=dims + [dims[0]],
            fill="toself",
            name=row.get("Company", f"#{i+1}"),
            line_color=pal[i % len(pal)],
            opacity=0.72,
        ))
    fig.update_layout(
        polar=dict(radialaxis=dict(visible=True, range=[0, 10])),
        showlegend=True, title=title, height=430,
        margin=dict(l=10, r=10, t=45, b=10),
    )
    return fig


def init_rows(key: str, n: int = 3):
    if key not in st.session_state:
        st.session_state[key] = [{"country": None, "value": 0.0} for _ in range(n)]


def add_row(key: str):
    st.session_state[key].append({"country": None, "value": 0.0})


def country_input_table(key: str, country_options: list) -> dict:
    """Render a dynamic country / weight input table and return {country: raw_weight}."""
    rows   = st.session_state[key]
    result = {}
    total  = sum(r["value"] for r in rows if r["country"] and r["value"] > 0)

    hcols = st.columns([3, 2, 1])
    hcols[0].markdown("**Country**")
    hcols[1].markdown("**Weight (%)**")
    hcols[2].markdown("**Share**")

    for i, row in enumerate(rows):
        c1, c2, c3 = st.columns([3, 2, 1])

        # safe index lookup
        idx = 0
        if row["country"] and row["country"] in country_options:
            idx = country_options.index(row["country"]) + 1

        country = c1.selectbox(
            "Country", options=[""] + country_options,
            index=idx, key=f"{key}_c_{i}",
            label_visibility="collapsed",
        )
        value = c2.number_input(
            "Weight", min_value=0.0, step=1.0,
            value=float(row["value"]), key=f"{key}_v_{i}",
            label_visibility="collapsed",
        )
        share_txt = f"{value / total * 100:.1f}%" if total > 0 and country and value > 0 else ""
        c3.markdown(
            f"<div style='padding-top:8px;color:#888;font-size:0.85rem'>{share_txt}</div>",
            unsafe_allow_html=True,
        )

        row["country"] = country or None
        row["value"]   = value
        if country and value > 0:
            result[country] = value

    if total > 0 and abs(total - 100) > 0.5:
        st.caption(f"⚠ Weights sum to {total:.1f} — will be auto-normalised to 100 %.")

    return result


# ═══════════════════════════════════════════════════════════════════════════
# APP
# ═══════════════════════════════════════════════════════════════════════════

country_df, country_risk, sector_df, sector_mult, bench_df, vol_mult = load_all()
country_options = sorted(country_risk.keys())
sector_options  = sorted(sector_df["sector"].unique())

st.title("🌐 Corporate Geopolitical Risk Index")
st.caption(
    "Methodology follows the CGRI Excel framework (2024 edition). "
    f"Volatility multiplier (2024): **{vol_mult:.4f}** — sourced from CBOE VIX via FRED."
)

tab_dash, tab_calc = st.tabs(["📊 Benchmark Dashboard", "🧮 Custom Calculator"])


# ─────────────────────────── TAB 1: BENCHMARK ───────────────────────────────
with tab_dash:
    st.subheader("25-Corporate Benchmark Portfolio (2024)")

    # ── Filters ─────────────────────────────────────────────────────────────
    f1, f2, f3 = st.columns([2, 2, 1])
    risk_filter   = f1.multiselect("Risk category",
                                   options=list(RISK_COLORS.keys()),
                                   default=list(RISK_COLORS.keys()))
    sector_filter = f2.multiselect("Sector",
                                   options=sorted(bench_df["Sector"].unique()),
                                   default=sorted(bench_df["Sector"].unique()))
    asc           = f3.checkbox("Sort ascending", value=False)

    view = bench_df[
        bench_df["Risk Category"].isin(risk_filter) &
        bench_df["Sector"].isin(sector_filter)
    ].sort_values("Final CGRI", ascending=asc)

    if view.empty:
        st.warning("No companies match the current filters.")
        st.stop()

    # ── Overall bar chart ────────────────────────────────────────────────────
    fig_bar = px.bar(
        view, x="Company", y="Final CGRI",
        color="Risk Category", color_discrete_map=RISK_COLORS,
        hover_data=["Sector"] + COMPONENT_KEYS + ["Sector Multiplier"],
        title="Final CGRI Score",
        height=420,
    )
    fig_bar.update_layout(xaxis_tickangle=-40, legend_title="Category")
    st.plotly_chart(fig_bar, use_container_width=True)

    # ── Component stacked bar ────────────────────────────────────────────────
    with st.expander("Weighted component breakdown (before multipliers)"):
        cb = view.copy()
        cb["HQ Risk (w)"]         = 0.15 * cb["HQ Risk"]
        cb["Revenue (w)"]          = 0.25 * cb["Revenue Exposure"]
        cb["Supply Chain (w)"]     = 0.25 * cb["Supply Chain"]
        cb["Facility (w)"]         = 0.15 * cb["Facility Risk"]
        cb["Financial (w)"]        = 0.20 * cb["Financial Exposure"]
        melted = cb[["Company",
                     "HQ Risk (w)", "Revenue (w)", "Supply Chain (w)",
                     "Facility (w)", "Financial (w)"]].melt(
            id_vars="Company", var_name="Component", value_name="Score"
        )
        fig_stack = px.bar(
            melted, x="Company", y="Score", color="Component",
            barmode="stack", height=380,
            title="Weighted pre-multiplier contribution",
        )
        fig_stack.update_layout(xaxis_tickangle=-40)
        st.plotly_chart(fig_stack, use_container_width=True)

    # ── Radar comparison ─────────────────────────────────────────────────────
    st.markdown("#### Radar comparison")
    sel_cos = st.multiselect(
        "Select up to 5 companies to compare",
        options=view["Company"].tolist(),
        default=view["Company"].tolist()[:3],
    )
    if sel_cos:
        radar_rows = (
            view[view["Company"].isin(sel_cos)]
            .rename(columns={"Supply Chain": "Supply Chain"})[
                ["Company"] + COMPONENT_KEYS
            ].to_dict("records")
        )
        st.plotly_chart(radar_chart(radar_rows, "Risk dimension comparison"), use_container_width=True)

    # ── Data table ───────────────────────────────────────────────────────────
    st.markdown("#### Full data table")
    disp_cols = ["Company", "Sector", "Risk Category", "Final CGRI"] + COMPONENT_KEYS + \
                ["Sector Multiplier", "Volatility Multiplier"]
    fmt = {
        "Final CGRI": "{:.2f}", "HQ Risk": "{:.2f}", "Revenue Exposure": "{:.2f}",
        "Supply Chain": "{:.2f}", "Facility Risk": "{:.2f}", "Financial Exposure": "{:.0f}",
        "Sector Multiplier": "{:.2f}", "Volatility Multiplier": "{:.4f}",
    }
    st.dataframe(view[disp_cols].reset_index(drop=True).style.format(fmt), use_container_width=True)

    csv_bytes = view[disp_cols].to_csv(index=False).encode()
    st.download_button("⬇ Download CSV", data=csv_bytes,
                       file_name="cgri_benchmark_2024.csv", mime="text/csv")


# ─────────────────────────── TAB 2: CALCULATOR ──────────────────────────────
with tab_calc:
    st.subheader("Custom Corporate CGRI Calculator")
    st.caption(
        "Enter country-level exposure weights (any unit — auto-normalised to 100%). "
        "All fields must have at least one country entry to be scored."
    )

    # ── Corporate metadata ─────────────────────────────────────────────────────
    r1c1, r1c2 = st.columns(2)
    company_name  = r1c1.text_input("Corporate name", value="My Corporate")
    hq_country    = r1c2.selectbox("HQ country", country_options)

    r2c1, r2c2 = st.columns(2)
    sector        = r2c1.selectbox("Sector (S&P Global category)", sector_options)
    net_debt_ebitda = r2c2.number_input(
        "Net Debt / EBITDA (2024)",
        value=1.0, step=0.1, format="%.2f",
        help="Used to derive the Financial Exposure score (2–10 scale)."
    )

    # Show derived financial score
    fin_preview = net_debt_to_financial_score(net_debt_ebitda)
    r2c2.caption(f"→ Financial Exposure score: **{fin_preview}** / 10")

    st.divider()

    # ── Exposure inputs ──────────────────────────────────────────────────────
    init_rows("rev");  init_rows("sup");  init_rows("supfac");  init_rows("fac")

    col_a, col_b = st.columns(2)

    with col_a:
        st.markdown("#### 1. Revenue by country")
        st.caption("% of total revenues generated in each country")
        rev_input = country_input_table("rev", country_options)
        if st.button("➕ Add country", key="add_rev"):
            add_row("rev"); st.rerun()

        st.markdown("#### 3. Supplier distribution by country")
        st.caption("% of suppliers located in each country")
        sup_input = country_input_table("sup", country_options)
        if st.button("➕ Add country", key="add_sup"):
            add_row("sup"); st.rerun()

    with col_b:
        st.markdown("#### 2. Facility sites by country")
        st.caption("% of owned/operated facility sites in each country")
        fac_input = country_input_table("fac", country_options)
        if st.button("➕ Add country", key="add_fac"):
            add_row("fac"); st.rerun()

        st.markdown("#### 4. Supplier facility distribution by country")
        st.caption("% of supplier facilities located in each country")
        supfac_input = country_input_table("supfac", country_options)
        if st.button("➕ Add country", key="add_supfac"):
            add_row("supfac"); st.rerun()

    st.divider()

    # ── Compute ──────────────────────────────────────────────────────────────
    if st.button("🔍  Compute CGRI", type="primary", use_container_width=True):
        try:
            result = compute_cgri(
                hq_country=hq_country,
                sector=sector,
                net_debt_ebitda=net_debt_ebitda,
                revenue_by_country=rev_input,
                suppliers_by_country=sup_input,
                supplier_facilities_by_country=supfac_input,
                facility_by_country=fac_input,
                country_risk=country_risk,
                sector_mult_lookup=sector_mult,
                volatility_mult=vol_mult,
                company_name=company_name,
            )

            final = result["final_cgri"]
            cat, col = risk_label(final)

            st.markdown("---")
            st.markdown(f"### Results for **{company_name}**")

            # ── Headline badge ───────────────────────────────────────────────
            bx, _, _ = st.columns([1, 1, 1])
            bx.markdown(
                f"""<div style='background:{col};padding:18px 24px;border-radius:10px;text-align:center'>
                <div style='font-size:2.4rem;font-weight:700;color:#fff'>{final:.2f}</div>
                <div style='font-size:1rem;color:#fff'>Final CGRI — <b>{cat} Risk</b></div>
                </div>""",
                unsafe_allow_html=True,
            )

            # ── Component metrics ────────────────────────────────────────────
            st.markdown("#### Component breakdown")
            m1, m2, m3, m4, m5 = st.columns(5)
            m1.metric("HQ Risk",            f"{result['hq_risk']:.2f}")
            m2.metric("Revenue Exposure",    f"{result['revenue_exposure']:.2f}",
                      delta=f"HHI {result['rev_hhi']:.2f} → ×{result['rev_hhi_sub']:.2f}",
                      delta_color="off")
            m3.metric("Supply Chain",        f"{result['supply_chain']:.2f}",
                      delta=f"HHI {result['sc_hhi_combined']:.2f} → ×{result['sc_hhi_sub']:.2f}",
                      delta_color="off")
            m4.metric("Facility Risk",       f"{result['facility_risk']:.2f}")
            m5.metric("Financial Exposure",  f"{result['financial_exposure']:.0f}",
                      delta=f"Net D/EBITDA {net_debt_ebitda:.2f}",
                      delta_color="off")

            m6, m7, _ = st.columns(3)
            m6.metric("Sector Multiplier",   f"×{result['sector_multiplier']:.2f}",
                      delta=sector, delta_color="off")
            m7.metric("Volatility Multiplier", f"×{result['volatility_multiplier']:.4f}",
                      delta="2024 VIX avg vs long-term", delta_color="off")

            # ── Supply chain detail ──────────────────────────────────────────
            with st.expander("Supply chain detail"):
                st.markdown(
                    f"""
| | Value |
|---|---|
| Suppliers component (Σ GRI × sup_share) | {result['sc_sup_component']:.4f} |
| Supplier-facilities component (Σ GRI × fac_share) | {result['sc_fac_component']:.4f} |
| Intermediate (0.5 × A × 0.5 × B) | {result['sc_intermediate']:.4f} |
| HHI suppliers | {result['sc_hhi_sup']:.4f} |
| HHI supplier facilities | {result['sc_hhi_fac']:.4f} |
| HHI combined (avg) | {result['sc_hhi_combined']:.4f} |
| HHI submultiplier | {result['sc_hhi_sub']:.2f} |
| **Supply Chain Exposure** | **{result['supply_chain']:.4f}** |
"""
                )

            # ── Radar + benchmark comparison ─────────────────────────────────
            ch1, ch2 = st.columns(2)

            with ch1:
                custom_radar = {
                    "Company":          company_name,
                    "HQ Risk":          result["hq_risk"],
                    "Revenue Exposure": result["revenue_exposure"],
                    "Supply Chain":     result["supply_chain"],
                    "Facility Risk":    result["facility_risk"],
                    "Financial Exposure": result["financial_exposure"],
                }
                st.plotly_chart(
                    radar_chart([custom_radar], f"{company_name} — Risk profile"),
                    use_container_width=True,
                )

            with ch2:
                custom_row = pd.DataFrame([{
                    "Company": f"▶ {company_name}",
                    "Final CGRI": final,
                    "Risk Category": cat,
                }])
                compare = pd.concat(
                    [bench_df[["Company", "Final CGRI", "Risk Category"]], custom_row],
                    ignore_index=True,
                ).sort_values("Final CGRI", ascending=False)
                fig_cmp = px.bar(
                    compare, x="Company", y="Final CGRI",
                    color="Risk Category", color_discrete_map=RISK_COLORS,
                    title="vs. benchmark portfolio",
                    height=380,
                )
                fig_cmp.update_layout(xaxis_tickangle=-40, showlegend=False)
                st.plotly_chart(fig_cmp, use_container_width=True)

            # ── Export ───────────────────────────────────────────────────────
            st.markdown("#### Export")
            e1, e2 = st.columns(2)

            summary = {
                "Company":                company_name,
                "HQ Country":             result["hq_country"],
                "Sector":                 result["sector"],
                "Net Debt/EBITDA":        result["net_debt_ebitda"],
                "HQ Risk":                round(result["hq_risk"], 4),
                "Revenue Exposure":       round(result["revenue_exposure"], 4),
                "Revenue HHI":            round(result["rev_hhi"], 4),
                "Revenue HHI Sub":        result["rev_hhi_sub"],
                "Supply Chain Exposure":  round(result["supply_chain"], 4),
                "SC Sup Component":       round(result["sc_sup_component"], 4),
                "SC Fac Component":       round(result["sc_fac_component"], 4),
                "SC Intermediate":        round(result["sc_intermediate"], 4),
                "SC HHI Combined":        round(result["sc_hhi_combined"], 4),
                "SC HHI Sub":             result["sc_hhi_sub"],
                "Facility Risk":          round(result["facility_risk"], 4),
                "Financial Exposure":     result["financial_exposure"],
                "Sector Multiplier":      result["sector_multiplier"],
                "Volatility Multiplier":  round(result["volatility_multiplier"], 4),
                "Final CGRI":             round(result["final_cgri"], 4),
                "Risk Category":          cat,
            }

            e1.download_button(
                "⬇ CSV (summary)",
                data=pd.DataFrame([summary]).to_csv(index=False).encode(),
                file_name=f"{company_name.replace(' ','_')}_cgri.csv",
                mime="text/csv",
            )

            export_json = {k: v for k, v in result.items()
                          if not isinstance(v, dict)}
            export_json["rev_shares"]     = result["rev_shares"]
            export_json["sup_shares"]     = result["sup_shares"]
            export_json["fac_sup_shares"] = result["fac_sup_shares"]
            export_json["fac_shares"]     = result["fac_shares"]
            e2.download_button(
                "⬇ JSON (full detail)",
                data=json.dumps(export_json, indent=2, default=str),
                file_name=f"{company_name.replace(' ','_')}_cgri.json",
                mime="application/json",
            )

        except Exception as exc:
            st.error(f"⚠ {exc}")
            st.exception(exc)


# ─────────────────────────── Methodology note ───────────────────────────────
with st.expander("ℹ Methodology & data sources"):
    st.markdown(r"""
### CGRI Formula (2024)

$$
\text{CGRI} = \bigl(
  0.15 \cdot \text{HQ Risk}
+ 0.25 \cdot \text{Revenue Exp.}
+ 0.25 \cdot \text{Supply Chain Exp.}
+ 0.15 \cdot \text{Facility Risk}
+ 0.20 \cdot \text{Financial Exp.}
\bigr)
\times \text{Sector Mult.} \times \text{Volatility Mult.}
$$

### Component formulas
| Component | Formula |
|---|---|
| **HQ Risk** | Country GRI of headquarters location |
| **Revenue Exposure** | $\Sigma(\text{GRI}_c \times \text{rev\_std}_c) \times \text{HHI\_sub}$ |
| **Supply Chain** | $0.5 \times C_{\text{sup}} \times 0.5 \times C_{\text{fac}} \times \text{HHI\_sub}$ |
| **Facility Risk** | $\Sigma(\text{GRI}_c \times \text{fac\_std}_c)$ — no HHI |
| **Financial Exposure** | Discrete score 2–10 from Net Debt / EBITDA |

### HHI Concentration Submultiplier
| HHI | Submultiplier | Interpretation |
|-----|-------------|----------------|
| < 0.15 | **0.90** | Highly diversified |
| 0.15 – 0.25 | **1.00** | Moderate |
| 0.25 – 0.40 | **1.10** | Somewhat concentrated |
| 0.40 – 0.60 | **1.25** | Concentrated |
| ≥ 0.60 | **1.50** | Highly concentrated |

### Risk categories
| Score | Category |
|---|---|
| < 3.5 | 🟢 Low |
| 3.5 – 5.0 | 🟡 Moderate |
| 5.0 – 6.5 | 🟠 High |
| ≥ 6.5 | 🔴 Very High |

### Financial Exposure scoring
| Net Debt / EBITDA | Score |
|---|---|
| < −1 | 2 |
| −1 to 0 | 3 |
| 0 to 1 | 4 |
| 1 to 2 | 5 |
| 2 to 3 | 6 |
| 3 to 4 | 7 |
| 4 to 5 | 8 |
| 5 to 6 | 9 |
| ≥ 6 | 10 |

### Data sources
- **Country GRI (2024)**: proprietary 147-country index
- **Sector multipliers**: S&P Global Ratings Industry Risk Assessment
- **Volatility multiplier**: CBOE VIX (FRED: VIXCLS) — 2024 annual avg vs. long-term avg of 19.87
""")
